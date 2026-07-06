"""
AtenLanaportal - Proof of Concept
Single-file Flask app implementing Lecturer/Student/SuperAdmin roles, token-based attendance with
location checks, grace period logic, device/IP verification, exports (Excel/PDF fallback), and inline HTML/CSS/JS.

Run: python index.py

Dependencies (recommended): Flask, pandas, openpyxl, fpdf
"""

from flask import Flask, g, render_template, request, redirect, url_for, session, jsonify, send_file, flash
import os
import re
import json
import threading
import requests
from urllib.parse import unquote
from bs4 import BeautifulSoup
import base64
import hashlib
from dotenv import load_dotenv

# Load environment variables from .env file if it exists
load_dotenv()
import random
import secrets
import string
import time
from datetime import datetime, timedelta, timezone
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from io import BytesIO
from flask_wtf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# Optional libs
try:
    import pandas as pd
except ImportError:
    pd = None

# try openpyxl directly for Excel export if pandas not present
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

# Database support for PostgreSQL
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None

# --- Configuration ---
_SETTINGS_CACHE = {}
_SETTINGS_CACHE_TIME = 0

def get_setting(key, default=None):
    global _SETTINGS_CACHE, _SETTINGS_CACHE_TIME
    now = time.time()
    # Cache for 30 seconds to reduce DB load on before_request hooks
    if not _SETTINGS_CACHE or (now - _SETTINGS_CACHE_TIME) > 30:
        try:
            # Note: query_db might not be available yet if called during early init,
            # but usually called in request context
            rows = query_db("SELECT key, value FROM settings")
            _SETTINGS_CACHE = {r['key']: r['value'] for r in rows}
            _SETTINGS_CACHE_TIME = now
        except Exception:
            return default
    return _SETTINGS_CACHE.get(key, default)

def get_database_url():
    # Priority: DATABASE_URL -> POSTGRES_URL (Vercel Integration)
    url = os.environ.get('DATABASE_URL') or os.environ.get('POSTGRES_URL')
    if url and url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return url

# Protocol Integrity: Enforce mandatory environment lattice in production
IS_PROD = os.environ.get('FLASK_ENV') == 'production'

DATABASE_URL = get_database_url()
# Vercel Protocol: Map ephemeral storage to /tmp for runtime operations
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', '/tmp/atenlana_uploads' if IS_PROD else 'static/img')
if not os.path.exists(UPLOAD_FOLDER):
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    except Exception as e:
        if not IS_PROD:
            print(f"Warning: Could not create upload folder {UPLOAD_FOLDER}: {e}")

# Use a more readable character set for tokens (exclude 0, 1, O, I)
TOKEN_CHARS = "23456789ABCDEFGHJKLMNPQRSTUVWXYZ"
DEFAULT_SESSION_DURATION_MIN = 20
DEFAULT_GRACE_MIN = 5
DEFAULT_RADIUS_M = 10  # meters

# System Config Lattice: Built-in defaults with environment overriding
SECRET_KEY = os.environ.get('SECRET_KEY', "atenlana-protocol-key-9988776655")
SUPERADMIN_PASS = os.environ.get('SUPERADMIN_PASS', '1234')
ADMIN_PASS = os.environ.get('ADMIN_PASS', 'admin123')
DEFAULT_PASS = "password"

FOUNDER_BIO = os.environ.get('FOUNDER_BIO', "CEO & Founder of AtenLana Grid Systems. Visionary Architect committed to high-fidelity presence validation.")
FOUNDER_EMAIL = os.environ.get('FOUNDER_EMAIL', 'omonire@atenlana.com')
FOUNDER_WHATSAPP = os.environ.get('FOUNDER_WHATSAPP', '+2348000000000')

OMONIRE_BIO = FOUNDER_BIO

# Robust path resolution for deployment
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app = Flask(__name__,
            static_folder=os.path.join(BASE_DIR, 'static'),
            template_folder=os.path.join(BASE_DIR, 'templates'))

print(f"Initializing AtenLana Portal (Base Dir: {BASE_DIR})")

app.config['SECRET_KEY'] = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# In modern high-security apps, we force secure cookies even in development if possible,
# but for local testing without HTTPS, we'll respect production flag.
if os.environ.get('FLASK_ENV') == 'production' or os.environ.get('FORCE_SECURE_COOKIES'):
    app.config['SESSION_COOKIE_SECURE'] = True

# Enable CSRF protection for forms (disabled by default in non-production for PoC/testing)
csrf = CSRFProtect()
# Initialize CSRF protection
csrf.init_app(app)

# Initialize Rate Limiter with dynamic toggle
def limiter_request_filter():
    try:
        status = get_setting('security_mode', 'off')
        return status == 'off'
    except Exception:
        return True # Default to filtered (disabled) if error

try:
    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["1000 per day", "100 per hour"],
        storage_uri="memory://"
    )
    limiter.request_filter(limiter_request_filter)
except Exception as e:
    app.logger.error(f"Limiter initialization failed: {e}")
    class NoOpLimiter:
        def limit(self, *args, **kwargs): return lambda f: f
    limiter = NoOpLimiter()

# --- Utilities ---
def security_log(event, details=None):
    ip = request.remote_addr or 'unknown'
    user_id = session.get('user_id', 'Guest')
    log_msg = f"SECURITY EVENT | {event} | IP: {ip} | User: {user_id} | Details: {details}"
    app.logger.warning(log_msg)

def log_action(user_id, action_type, target_table, target_id, old_data=None, new_data=None):
    # Security: Scrub sensitive fields before logging in new_data (the incoming change)
    # but preserve old_data for restoration (UNDO). Admins are trusted to see logs.
    def scrub(data):
        if not data: return data
        if isinstance(data, dict):
            return {k: ('***SCRUBBED***' if 'password' in k.lower() or 'hash' in k.lower() or 'security_a' in k.lower() else v)
                    for k, v in data.items()}
        return data

    try:
        commit_db("INSERT INTO action_history (user_id, action_type, target_table, target_id, old_data, new_data) VALUES (?,?,?,?,?,?)",
                  (user_id, action_type, target_table, target_id,
                   json.dumps(old_data, default=str) if old_data is not None else None,
                   json.dumps(scrub(new_data), default=str) if new_data is not None else None))
    except Exception as e:
        app.logger.error(f"Failed to log action: {e}")

import smtplib
from email.mime.text import MIMEText

def send_attendance_receipt(student_email, student_name, course_name, timestamp, session_id):
    if not student_email:
        return
    try:
        smtp_server = os.environ.get('MAIL_SERVER', '')
        smtp_port = int(os.environ.get('MAIL_PORT', '587'))
        smtp_user = os.environ.get('MAIL_USERNAME', '')
        smtp_pass = os.environ.get('MAIL_PASSWORD', '')
        if not smtp_server or not smtp_user:
            app.logger.info(f"Email receipt skipped for {student_email}: mail not configured")
            return
        msg = MIMEText(f"""Hello {student_name},

Your attendance has been confirmed for {course_name}.

• Session ID: {session_id}
• Time: {timestamp}
• Status: Present

Keep attending to earn rewards and maintain your streak!

- AtenLana Portal""")
        msg['Subject'] = f"Attendance Confirmed - {course_name}"
        msg['From'] = smtp_user
        msg['To'] = student_email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        app.logger.info(f"Attendance receipt sent to {student_email}")
    except Exception as e:
        app.logger.error(f"Failed to send attendance receipt: {e}")

def send_session_notification(student_email, student_name, course_name, token, lecturer_name):
    if not student_email:
        return
    try:
        smtp_server = os.environ.get('MAIL_SERVER', '')
        smtp_port = int(os.environ.get('MAIL_PORT', '587'))
        smtp_user = os.environ.get('MAIL_USERNAME', '')
        smtp_pass = os.environ.get('MAIL_PASSWORD', '')
        if not smtp_server or not smtp_user:
            return
        msg = MIMEText(f"""Hello {student_name},

A new attendance session has been created by {lecturer_name}.

• Course: {course_name}
• Session Code: {token}
• Status: Active

Open your AtenLana Portal dashboard and enter this code to mark your presence.

- AtenLana Portal""")
        msg['Subject'] = f"New Session: {course_name} - Code: {token}"
        msg['From'] = smtp_user
        msg['To'] = student_email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
    except Exception as e:
        app.logger.error(f"Failed to send session notification: {e}")

_db_local = threading.local()

def connect_to_db():
    db_url = get_database_url()
    if not db_url:
        app.logger.error("DATABASE_URL environment variable is not set.")
        return None
    if psycopg2 is None:
        app.logger.error("psycopg2 (or psycopg2-binary) is not installed.")
        return None

    max_retries = 2 # Reduced for faster serverless boot
    last_err = None
    for attempt in range(max_retries):
        try:
            # Use a short timeout to prevent Vercel boot timeouts
            conn = psycopg2.connect(db_url, connect_timeout=3)
            # Integrity: Set a statement timeout to prevent hanging the serverless node
            with conn.cursor() as cur:
                cur.execute("SET statement_timeout = 25000") # 25 seconds
            return conn
        except Exception as e:
            last_err = e
            app.logger.warning(f"Database connection attempt {attempt+1} failed: {e}")
            if attempt < max_retries - 1:
                time.sleep(0.5)

    security_log("DATABASE_CONNECTION_FAILURE", str(last_err))
    # Non-fatal during module load / early startup to let Vercel finish deployment check
    if not has_request_context():
        app.logger.error(f"PostgreSQL Connection failed at boot: {last_err}")
        return None
    raise RuntimeError(f"Could not connect to database after {max_retries} attempts. Error: {last_err}")

from flask import has_request_context

def get_db():
    if has_request_context():
        db = getattr(g, '_database', None)
        if db is None:
            db = connect_to_db()
            if db:
                g._database = db
            else:
                raise RuntimeError("Database connection failed.")
        return db
    else:
        db = getattr(_db_local, 'db', None)
        if db is None:
            db = connect_to_db()
            if db:
                _db_local.db = db
            else:
                # Outside request context, we might not want to raise a fatal error immediately
                return None
        return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def parse_dt(val):
    """Helper to parse datetime from either string or datetime object (PostgreSQL returns objects).
    Ensures returning a timezone-aware UTC datetime.
    """
    if not val:
        return None
    dt = None
    if isinstance(val, datetime):
        dt = val
    else:
        try:
            dt = datetime.fromisoformat(val.replace('Z', '+00:00'))
        except (ValueError, TypeError):
            return None

    if dt and dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt

def query_db(query, args=(), one=False):
    db = get_db()
    # Convert SQLite placeholders to PostgreSQL safely
    query = re.sub(r"\?(?=(?:[^']*'[^']*')*[^']*$)", "%s", query)
    with db.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(query, args)
        rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv

def commit_db(query, args=()):
    db = get_db()
    # Convert SQLite placeholders and INSERT OR IGNORE to PostgreSQL safely
    query = re.sub(r"\?(?=(?:[^']*'[^']*')*[^']*$)", "%s", query)
    if 'INSERT OR IGNORE' in query.upper():
        query = re.sub(r'INSERT OR IGNORE', 'INSERT', query, flags=re.IGNORECASE)
        if 'ON CONFLICT' not in query.upper():
            query += ' ON CONFLICT DO NOTHING'
    with db.cursor() as cur:
        cur.execute(query, args)
        db.commit()

def init_db():
    """Create tables and seed minimal users (PostgreSQL only)."""
    db = None
    try:
        db = get_db()
    except Exception as e:
        app.logger.warning(f"Database initialization deferred: {e}")
        return

    if not db:
        app.logger.warning("Database initialization deferred: No connection available.")
        return

    # Check if already initialized to skip heavy migration logic
    try:
        with db.cursor() as cur:
            cur.execute("SELECT 1 FROM users LIMIT 1")
            # Ensure ub_id exists as a minimum schema check for modern version
            cur.execute("SELECT ub_id FROM users LIMIT 1")

            # Check if superadmin exists to determine if seeding is done
            sa_exists = query_db("SELECT id FROM users WHERE role = 'superadmin' LIMIT 1", one=True)
            if sa_exists:
                return # Fast path: Already initialized and seeded
    except Exception:
        if db: db.rollback()
        # If column or table missing, we continue to migrations

    def safe_execute(query, args=(), commit=True):
        """Execute query in a new transaction/cursor, robustly handling errors."""
        try:
            # Convert SQLite placeholders
            query_psql = re.sub(r"\?(?=(?:[^']*'[^']*')*[^']*$)", "%s", query)
            with db.cursor() as cur:
                cur.execute(query_psql, args)
                if commit:
                    db.commit()
                return True
        except Exception as e:
            app.logger.error(f"init_db migration error: {e}")
            db.rollback()
            return False

    # 1. Base Tables Creation (Individual transactions for robustness)
    tables = [
        "CREATE TABLE IF NOT EXISTS users (id SERIAL PRIMARY KEY, role TEXT NOT NULL, password_hash TEXT NOT NULL, active INTEGER DEFAULT 1)",
        "CREATE TABLE IF NOT EXISTS sessions (id SERIAL PRIMARY KEY, token TEXT UNIQUE, created_at TIMESTAMP DEFAULT NOW(), active INTEGER DEFAULT 1)",
        "CREATE TABLE IF NOT EXISTS attendance (id SERIAL PRIMARY KEY, session_id INTEGER, student_id INTEGER, UNIQUE(session_id, student_id))",
        "CREATE TABLE IF NOT EXISTS locations (id SERIAL PRIMARY KEY)",
        "CREATE TABLE IF NOT EXISTS courses (id SERIAL PRIMARY KEY, name TEXT UNIQUE NOT NULL)",
        "CREATE TABLE IF NOT EXISTS faculties (id SERIAL PRIMARY KEY, name TEXT UNIQUE NOT NULL)",
        "CREATE TABLE IF NOT EXISTS departments (id SERIAL PRIMARY KEY)",
        "CREATE TABLE IF NOT EXISTS lecturer_courses (id SERIAL PRIMARY KEY)",
        "CREATE TABLE IF NOT EXISTS team_roles (id SERIAL PRIMARY KEY, name TEXT UNIQUE NOT NULL)",
        "CREATE TABLE IF NOT EXISTS team_members (id SERIAL PRIMARY KEY, name TEXT NOT NULL)",
        "CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)",
        "CREATE TABLE IF NOT EXISTS announcements (id SERIAL PRIMARY KEY, sender_id INTEGER, target_type TEXT, target_id INTEGER, message TEXT, priority TEXT DEFAULT 'normal', created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS student_cards (id SERIAL PRIMARY KEY, student_id INTEGER UNIQUE, card_data TEXT, file_type TEXT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS blocked_ips (ip TEXT PRIMARY KEY, reason TEXT, created_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS failed_logins (ip TEXT, username TEXT, attempted_at TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS action_history (id SERIAL PRIMARY KEY, user_id INTEGER, action_type TEXT, target_table TEXT, target_id INTEGER, old_data TEXT, new_data TEXT, timestamp TIMESTAMP DEFAULT NOW(), reverted INTEGER DEFAULT 0)",
        "CREATE TABLE IF NOT EXISTS traffic_intelligence (id SERIAL PRIMARY KEY, user_id INTEGER, ip TEXT, ua TEXT, lat REAL, lon REAL, url TEXT, device_info TEXT, timestamp TIMESTAMP DEFAULT NOW())",
        "CREATE TABLE IF NOT EXISTS security_alerts (id SERIAL PRIMARY KEY, user_id INTEGER, type TEXT, details TEXT, ip TEXT, created_at TIMESTAMP DEFAULT NOW(), status TEXT DEFAULT 'open')"
    ]
    # Perform all creations in a single transaction to reduce latency
    try:
        with db.cursor() as cur:
            for sql in tables:
                cur.execute(sql)
        db.commit()
    except Exception as e:
        app.logger.error(f"Bulk table creation failed: {e}")
        db.rollback()

    # 2. Sequential Migrations
    # Users
    for col, ctype in [('role', 'TEXT NOT NULL'), ('username', 'TEXT'), ('password_hash', 'TEXT NOT NULL'),
                       ('first_name', 'TEXT'), ('middle_name', 'TEXT'), ('last_name', 'TEXT'),
                       ('ub_id', 'TEXT'), ('security_q1', 'TEXT'), ('security_a1', 'TEXT'),
                       ('security_q2', 'TEXT'), ('security_a2', 'TEXT'), ('current_ip', 'TEXT'),
                       ('last_activity', 'TIMESTAMP'), ('active', 'INTEGER DEFAULT 1'),
                       ('department_id', 'INTEGER'), ('matric', 'TEXT'),
                       ('biometric_data', 'TEXT'), ('faculty_id', 'INTEGER'),
                       ('assigned_dept_id', 'INTEGER'), ('github_token', 'TEXT'),
                       ('vercel_token', 'TEXT'), ('vercel_project_id', 'TEXT'),
                        ('blocked_until', 'TIMESTAMP'),
                        ('reward_points', 'INTEGER DEFAULT 0'),
                        ('streak', 'INTEGER DEFAULT 0'),
                        ('last_attendance_date', 'DATE'),
                        ('wallet_address', 'TEXT'),
                        ('email', 'TEXT')]:
        safe_execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Sessions
    for col, ctype in [('token', 'TEXT UNIQUE'), ('lecturer_id', 'INTEGER'), ('course_id', 'INTEGER'),
                       ('course_name', 'TEXT'), ('host', 'TEXT'), ('start_time', 'TIMESTAMP'),
                       ('duration_min', 'INTEGER'), ('grace_min', 'INTEGER'), ('latitude', 'REAL'),
                       ('longitude', 'REAL'), ('radius_m', 'REAL'), ('attendance_mode', "TEXT DEFAULT 'both'"),
                       ('strict_mode', 'INTEGER DEFAULT 1'), ('active', 'INTEGER DEFAULT 1')]:
        safe_execute(f"ALTER TABLE sessions ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Attendance
    for col, ctype in [('session_id', 'INTEGER'), ('student_id', 'INTEGER'), ('timestamp', 'TIMESTAMP'),
                       ('ip', 'TEXT'), ('device_fp', 'TEXT'), ('lat', 'REAL'), ('lon', 'REAL'),
                       ('status', 'TEXT'), ('first_name', 'TEXT'), ('middle_name', 'TEXT'), ('last_name', 'TEXT')]:
        safe_execute(f"ALTER TABLE attendance ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Locations
    for col, ctype in [('lecturer_id', 'INTEGER'), ('name', 'TEXT'), ('latitude', 'REAL'), ('longitude', 'REAL')]:
        safe_execute(f"ALTER TABLE locations ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Faculties
    for col, ctype in [('code', 'TEXT UNIQUE')]:
        safe_execute(f"ALTER TABLE faculties ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Departments
    for col, ctype in [('faculty_id', 'INTEGER'), ('name', 'TEXT NOT NULL'), ('code', 'TEXT')]:
        safe_execute(f"ALTER TABLE departments ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Courses
    for col, ctype in [('code', 'TEXT'), ('title', 'TEXT')]:
        safe_execute(f"ALTER TABLE courses ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Lecturer Courses
    for col, ctype in [('lecturer_id', 'INTEGER'), ('course_id', 'INTEGER')]:
        safe_execute(f"ALTER TABLE lecturer_courses ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Team Members - handle legacy and fix constraints
    safe_execute("""
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='team_members' AND column_name='role') THEN
                ALTER TABLE team_members ALTER COLUMN role DROP NOT NULL;
            END IF;
        END;
        $$;
    """)

    safe_execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_dept_code_faculty') THEN
                ALTER TABLE departments ADD CONSTRAINT unique_dept_code_faculty UNIQUE(faculty_id, code);
            END IF;
        END;
        $$;
    """)
    for col, ctype in [('role_id', 'INTEGER'), ('bio', 'TEXT'), ('image', 'TEXT'), ('visible', 'INTEGER DEFAULT 1'),
                       ('email', 'TEXT'), ('whatsapp', 'TEXT')]:
        safe_execute(f"ALTER TABLE team_members ADD COLUMN IF NOT EXISTS {col} {ctype}")

    # Resolve NotNullViolation for unknown columns in team_members
    try:
        with db.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'team_members' AND is_nullable = 'NO'
                AND column_name NOT IN ('id', 'name') AND column_default IS NULL
            """)
            for row in cur.fetchall():
                safe_execute(f"ALTER TABLE team_members ALTER COLUMN {row['column_name']} DROP NOT NULL")
    except Exception:
        db.rollback()

    # Constraints & Indices
    # Use DO blocks for robust idempotency on PostgreSQL
    safe_execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_class WHERE relname = 'unique_session_student') AND
               NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_session_student') THEN
                ALTER TABLE attendance ADD CONSTRAINT unique_session_student UNIQUE(session_id, student_id);
            END IF;
            -- High-Level Strictness: Ensure unique identity lattice
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_username') THEN
                ALTER TABLE users ADD CONSTRAINT unique_username UNIQUE(username);
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_ub_id') THEN
                ALTER TABLE users ADD CONSTRAINT unique_ub_id UNIQUE(ub_id);
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_matric') THEN
                ALTER TABLE users ADD CONSTRAINT unique_matric UNIQUE(matric);
            END IF;
        END;
        $$;
    """)
    safe_execute("CREATE INDEX IF NOT EXISTS idx_attendance_student_id ON attendance(student_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_attendance_session_device ON attendance(session_id, device_fp)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_sessions_lecturer_id ON sessions(lecturer_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_users_role ON users(role)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_users_ub_id ON users(ub_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_users_matric ON users(matric)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_announcements_target ON announcements(target_type, target_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_failed_logins_ip ON failed_logins(ip, attempted_at)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_student_cards_user ON student_cards(student_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_users_faculty ON users(faculty_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_users_assigned_dept ON users(assigned_dept_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_action_history_user ON action_history(user_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_action_history_time ON action_history(timestamp)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_traffic_intel_time ON traffic_intelligence(timestamp)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_security_alerts_user ON security_alerts(user_id)")
    safe_execute("CREATE INDEX IF NOT EXISTS idx_security_alerts_time ON security_alerts(created_at)")

    # Unique constraint for departments to prevent duplicates during sync
    safe_execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'unique_dept_name_faculty') THEN
                ALTER TABLE departments ADD CONSTRAINT unique_dept_name_faculty UNIQUE(faculty_id, name);
            END IF;
        END;
        $$;
    """)

    # Add priority column to announcements if missing (Migration)
    safe_execute("ALTER TABLE announcements ADD COLUMN IF NOT EXISTS priority TEXT DEFAULT 'normal'")

    # Conditional Foreign Key addition to avoid "already exists" errors
    safe_execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'team_members_role_id_fkey') THEN
                ALTER TABLE team_members ADD CONSTRAINT team_members_role_id_fkey FOREIGN KEY (role_id) REFERENCES team_roles(id);
            END IF;
        END;
        $$;
    """)

    # Seed website status if missing
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('website_status', 'on'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('security_mode', 'off'))
        # High-Level Strictness: Global Integrity Protocol enabled by default for all nodes
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('strict_integrity_mode', 'on'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('primary_color', '#ffba08'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('secondary_color', '#ffba08'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('github_access_for_admin', 'off'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('student_cost', '10000'))
    safe_execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO NOTHING", ('billing_period', 'session'))

    # If security mode is off, clear blocked IPs on startup
    sm = query_db("SELECT value FROM settings WHERE key = 'security_mode'", one=True)
    if not sm or sm['value'] == 'off':
        safe_execute("DELETE FROM blocked_ips")
        safe_execute("DELETE FROM failed_logins")

    # Seed superadmin
    try:
        sa_exists = query_db("SELECT id FROM users WHERE role = 'superadmin'", one=True)
        if not sa_exists:
            commit_db("INSERT INTO users (role, username, password_hash, first_name, active) VALUES (?,?,?,?,?)",
                      ('superadmin', 'superadmin', generate_password_hash(SUPERADMIN_PASS), 'Super', 1))
    except Exception as e:
        app.logger.error(f"Error seeding superadmin: {e}")

    # Seed default team member if empty
    try:
        # Ensure 'CEO & Founder' role exists
        commit_db('INSERT INTO team_roles (name) VALUES (?) ON CONFLICT DO NOTHING', ('CEO & Founder',))
        r_row = query_db('SELECT id FROM team_roles WHERE name = ?', ('CEO & Founder',), one=True)

        m_check = query_db('SELECT id FROM team_members WHERE name = ?', ('Omonire O. Great',), one=True)
        if not m_check:
            if r_row:
                # Default avatar SVG
                default_img = "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iODAwIiBoZWlnaHQ9IjgwMCIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cmVjdCB3aWR0aD0iMTAwJSIgaGVpZ2h0PSIxMDAlIiBmaWxsPSIjZjNmNGY2Ii8+PGNpcmNsZSBjeD0iNDAwIiBjeT0iMzAwIiByPSIxNTAiIGZpbGw9IiNkMWQ1ZGIiLz48cGF0aCBkPSJNNTAsODAwIEM1MCw2MDAgMjAwLDUwMCA0MDAsNTAwIEM2MDAsNTAwIDc1MCw2MDAgNzUwLDgwMCIgZmlsbD0iI2QxZDVkYiIvPjwvc3ZnPg=="
                commit_db('INSERT INTO team_members (role_id, name, bio, image, visible) VALUES (?,?,?,?,?)',
                          (r_row['id'], 'Omonire O. Great', OMONIRE_BIO, default_img, 1))
        else:
            # Update existing Omonire's bio to the new one
            commit_db('UPDATE team_members SET bio = ? WHERE name = ?', (OMONIRE_BIO, 'Omonire O. Great'))
    except Exception as e:
        app.logger.error(f"Error seeding Omonire O. Great: {e}")

    # Identity Matrix: Robust auto-seeding of all mandatory clusters
    try:
        # Check if Admin and Lecturer clusters exist
        admin_check = query_db("SELECT id FROM users WHERE role = 'admin' LIMIT 1", one=True)
        lecturer_check = query_db("SELECT id FROM users WHERE role = 'lecturer' LIMIT 1", one=True)
        student_check = query_db("SELECT id FROM users WHERE role = 'student' LIMIT 1", one=True)

        if not admin_check:
             commit_db('INSERT INTO users (role, username, password_hash, first_name, active) VALUES (?,?,?,?,?)',
                       ('admin', 'admin', generate_password_hash(ADMIN_PASS), 'System Admin', 1))

        if not lecturer_check or not student_check:
            print('Compulsory seeding: Re-establishing identity clusters...')
            # Sample courses
            sample_courses = ['CSC 101', 'MTH 201', 'PHY 102', 'CHM 101', 'GST 111']
            for cn in sample_courses:
                commit_db('INSERT INTO courses (name) VALUES (?) ON CONFLICT (name) DO NOTHING', (cn,))

            db = get_db()
            with db.cursor() as cur:
                if not lecturer_check:
                    cur.execute('INSERT INTO users (role, username, password_hash, first_name, middle_name, last_name, ub_id) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (username) DO NOTHING RETURNING id',
                                     ('lecturer', 'lecturer1', generate_password_hash(DEFAULT_PASS), 'Ada', '', 'Lecturer', 'UB_L_101'))
                    res = cur.fetchone()
                    lecturer_id = res[0] if res else None
                else:
                    cur.execute('SELECT id FROM users WHERE role = %s LIMIT 1', ('lecturer',))
                    lecturer_id = cur.fetchone()[0]

                if not student_check:
                    cur.execute('INSERT INTO users (role, username, password_hash, first_name, middle_name, last_name, ub_id, matric) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (ub_id) DO NOTHING RETURNING id',
                                     ('student', 'student1', generate_password_hash(DEFAULT_PASS), 'John', 'M', 'Doe', 'UB1001', 'MTH/2023/001'))
                    cur.execute('INSERT INTO users (role, username, password_hash, first_name, middle_name, last_name, ub_id, matric) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (ub_id) DO NOTHING RETURNING id',
                                     ('student', 'student2', generate_password_hash(DEFAULT_PASS), 'Jane', '', 'Smith', 'UB1002', 'CSC/2023/002'))
                db.commit()

            # Link lecturer to first course if new
            c = query_db('SELECT id, name FROM courses WHERE name = ?', ('CSC 101',), one=True)
            if c and lecturer_id:
                commit_db('INSERT INTO lecturer_courses (lecturer_id, course_id) VALUES (?,?) ON CONFLICT DO NOTHING', (lecturer_id, c['id']))
    except Exception as e:
        app.logger.error(f"Seeding Identity Matrix failed: {e}")


# Token generator
def generate_token(length=8):
    # Use secure randomness for tokens
    return ''.join(secrets.choice(TOKEN_CHARS) for _ in range(length))

# Use helpers for location and device fingerprinting
from location import haversine, geocode
from solana_client import mint_token, mint_badge, record_attendance, get_assets
from auth_helpers import device_fp


def evaluate_session_for_student(s, student_id, lat, lon, ip, ua, device_id=None):
    """Return structured info about whether a student can mark for a session.
    Fields: distance_m, within_radius, status ('active'|'ended'|'grace'),
    time_until_end (s), grace_seconds, already_marked, device_conflict, other_device_used, can_mark, message
    """
    try:
        start = parse_dt(s.get('start_time'))
        if not start:
            raise ValueError("Invalid session start time")

        duration = s.get('duration_min') or DEFAULT_SESSION_DURATION_MIN
        end = start + timedelta(minutes=duration)
        now = datetime.now(timezone.utc)
        mode = s.get('attendance_mode', 'both')
        strict = s.get('strict_mode', 1)

        distance_m = None
        within_radius = False

        # In 'location' or 'both' mode, we MUST have valid coordinates
        if mode in ('location', 'both'):
            if lat is not None and lon is not None and s.get('latitude') is not None and s.get('longitude') is not None:
                try:
                    distance_m = haversine(lat, lon, s['latitude'], s['longitude'])
                    within_radius = distance_m <= (s.get('radius_m') or DEFAULT_RADIUS_M)
                except Exception:
                    distance_m = None
                    within_radius = False
            else:
                within_radius = False
        else:
            # Token mode bypasses GPS radius check
            within_radius = True

        already_marked = query_db('SELECT * FROM attendance WHERE session_id = ? AND student_id = ?', (s['id'], student_id), one=True) is not None
        fp = device_fp(ip, ua, device_id)

        # High-Level Strictness: Global Integrity Protocol
        strict_integrity = get_setting('strict_integrity_mode', 'off') == 'on'
        device_conflict = False
        other_device_used = False

        if strict_integrity:
            # ONE Student Per Device: Has this device already been used by someone else for this session?
            conflict_row = query_db('SELECT student_id FROM attendance WHERE session_id = ? AND device_fp = ? AND student_id != ?',
                                    (s['id'], fp, student_id), one=True)
            if conflict_row:
                device_conflict = True

            # ONE Device Per Student: Has this student already used a different device for this session?
            other_row = query_db('SELECT device_fp FROM attendance WHERE session_id = ? AND student_id = ? AND device_fp != ?',
                                 (s['id'], student_id, fp), one=True)
            if other_row:
                other_device_used = True

        status = 'active' if now <= end else 'ended'
        time_until_end = int((end - now).total_seconds()) if now <= end else 0
        grace_seconds = None
        in_grace = False
        if now > end:
            grace_end = end + timedelta(minutes=s.get('grace_min') or 0)
            if now <= grace_end:
                in_grace = True
                grace_seconds = int((grace_end - now).total_seconds())
                status = 'grace'

        # Determine can_mark rules
        can_mark = False
        message = 'Time Expired'

        # Check if session is terminated
        is_active = (s.get('active') if 'active' in s.keys() else 1) == 1

        if already_marked:
            message = 'Already marked attendance for this session.'
            can_mark = False
        elif not is_active:
            message = 'This session has been terminated by the lecturer.'
            can_mark = False
        elif device_conflict:
            message = 'ONE STUDENT PER DEVICE: This device is already linked to another student for this session.'
            can_mark = False
        elif other_device_used:
            message = 'ONE DEVICE PER STUDENT: Your identity is already linked to a different device for this session.'
            can_mark = False
        else:
            if status == 'active' and within_radius:
                can_mark = True
                message = 'Attendance Approved (Active Period)'
            elif status == 'grace':
                # In grace period, we allow marking regardless of radius for better reliability
                # UNLESS Strict Integrity is enabled
                if strict_integrity and not within_radius and mode in ('location', 'both'):
                    can_mark = False
                    message = 'Grace period active but you are outside classroom radius (Strict Mode).'
                else:
                    can_mark = True
                    message = 'Attendance Approved (Grace Period)'
            else:
                # Provide better explanation
                if status == 'active' and not within_radius:
                    message = 'Session active but you are outside classroom radius. Please move closer.'
                else:
                    message = 'Session time has expired.'

        # Ensure start_time is a string for JSON/Template compatibility
        st = s.get('start_time')
        if isinstance(st, datetime):
            st = st.strftime('%Y-%m-%d %H:%M:%S')

        return {
            'session_id': s['id'],
            'course_name': s.get('course_name'),
            'start_time': st,
            'duration_min': duration,
            'grace_min': s.get('grace_min'),
            'distance_m': distance_m,
            'within_radius': within_radius,
            'status': status,
            'time_until_end': time_until_end,
            'grace_seconds': grace_seconds,
            'already_marked': already_marked,
            'device_conflict': device_conflict,
            'other_device_used': other_device_used,
            'can_mark': can_mark,
            'message': message
        }
    except Exception as e:
        app.logger.error(f"Error in evaluate_session_for_student (Session {s.get('id')}): {e}", exc_info=True)
        return {
            'session_id': s.get('id'),
            'can_mark': False,
            'message': f'Server evaluation failed: {str(e)[:50]}'
        }

# Initialize database on first request to prevent boot timeouts
_DB_INITIALIZED = False

@app.before_request
def auto_init_db():
    global _DB_INITIALIZED
    if not _DB_INITIALIZED:
        # Avoid running during static file requests or health checks
        if request.endpoint and ('static' in request.endpoint or 'health' in request.endpoint):
            return

        try:
            init_db()
            _DB_INITIALIZED = True
        except Exception as e:
            app.logger.error(f"Lazy database initialization failed: {e}")

# Authentication helpers - Toggleable via security_mode
@app.before_request
def security_scan():
    if request.endpoint and ('static' in request.endpoint or request.endpoint == 'health_check'):
        return

    try:
        if get_setting('security_mode', 'off') == 'off':
            return
    except Exception:
        return

    # Security: More robust WAF patterns
    patterns = [
        r"<script", r"javascript:", r"onerror=", r"onload=",
        r"UNION\s+SELECT", r"DROP\s+TABLE", r"OR\s+'?1'?='?1'?", r"WAITFOR\s+DELAY",
        r"exec\s+\(", r"cast\s+\(", r"declare\s+", r"truncate\s+"
    ]

    request_data = ""
    if request.method == 'POST':
        try:
            if request.is_json:
                request_data = json.dumps(request.get_json(silent=True) or "")
            else:
                request_data = str(request.form)
        except Exception:
            pass
    request_data += str(request.args)

    for pattern in patterns:
        try:
            if re.search(pattern, request_data, re.IGNORECASE):
                ip = request.remote_addr or 'unknown'
                security_log("WAF_TRIGGER", f"Pattern: {pattern}")
                commit_db("INSERT INTO blocked_ips (ip, reason) VALUES (?,?) ON CONFLICT (ip) DO NOTHING", (ip, f"WAF Trigger: {pattern}"))
                return "Malicious request detected. Your IP has been flagged.", 400
        except Exception as e:
            app.logger.error(f"WAF error: {e}")

@app.before_request
def firewall_check():
    if request.endpoint and ('static' in request.endpoint or request.endpoint == 'health_check'):
        return

    try:
        if get_setting('security_mode', 'off') == 'off':
            return
    except Exception:
        return

    ip = request.remote_addr or 'unknown'
    try:
        blocked = query_db("SELECT ip FROM blocked_ips WHERE ip = ?", (ip,), one=True)
        if blocked:
            return "Your IP address has been flagged for suspicious activity and is temporarily blocked.", 403
    except Exception:
        pass

@app.before_request
def check_website_status():
    # Only check if not a static file, not the login route, and not the superadmin toggling it
    if request.endpoint and 'static' not in request.endpoint and request.endpoint not in ['login', 'logout', 'index', 'health_check']:
        try:
            if get_setting('website_status', 'on') == 'off':
                # Allow superadmin to bypass
                user_id = session.get('user_id')
                if user_id:
                    user = query_db('SELECT role FROM users WHERE id = ?', (user_id,), one=True)
                    if user and user['role'] == 'superadmin':
                        return

                # If website is off and user is not superadmin, show maintenance (or redirect to index with message)
                if request.endpoint not in ['index', 'portal_hub']:
                    # Protocol Observation: Flash message causes session write, which might be slow on cold starts
                    flash('The portal is currently under maintenance.')
                    return redirect(url_for('index'))
        except Exception:
            pass

@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        try:
            g.user = query_db('SELECT * FROM users WHERE id = ?', (user_id,), one=True)
            if g.user:
                # Security Enforcement: Check for temporary blocks
                if g.user.get('blocked_until'):
                    blocked_until = parse_dt(g.user['blocked_until'])
                    if blocked_until and blocked_until > datetime.now(timezone.utc):
                        diff = blocked_until - datetime.now(timezone.utc)
                        hours, remainder = divmod(int(diff.total_seconds()), 3600)
                        minutes, _ = divmod(remainder, 60)
                        flash(f"Your account is temporarily suspended for security reasons. Remaining time: {hours}h {minutes}m.")
                        session.clear()
                        g.user = None
                        return

                # Update last activity only once every 5 minutes to reduce DB load
                last_update = session.get('last_activity_update')
                now = time.time()
                if last_update is None or (now - last_update) > 300:
                    commit_db('UPDATE users SET last_activity = ? WHERE id = ?', (datetime.now(timezone.utc), user_id))
                    session['last_activity_update'] = now
        except Exception:
            g.user = None

@app.context_processor
def inject_global_data():
    # Inject team members for the footer carousel and settings (logo/favicon)
    members = get_team_members_with_fallback(is_admin=False)
    settings = {}
    try:
        s_rows = query_db('SELECT * FROM settings')
        for r in s_rows:
            settings[r['key']] = r['value']
    except Exception:
        pass

    # Use get_setting to trigger/use cache for site_settings
    # This ensures site_settings is available in all templates
    get_setting('website_status') # Trigger refresh if needed

    # Filter out sensitive or obsolete settings from global context
    # Explicitly exclude chat settings as the feature has been removed
    excluded_keys = {'allow_chat', 'chat_persistence', 'chat_notifications'}
    filtered_settings = {k: v for k, v in _SETTINGS_CACHE.items() if k not in excluded_keys}

    return dict(team_members=members, site_settings=filtered_settings)

@app.after_request
def set_security_headers(response):
    # Advanced security headers
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'

    # Content Security Policy (Strict)
    # Allows CDNs used in the project (Tailwind, Google Fonts, Chart.js, Lucide, Anime.js)
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com https://cdn.jsdelivr.net https://unpkg.com https://cdnjs.cloudflare.com https://vercel.live https://*.vercel.live; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.tailwindcss.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://*.vercel.live https://cdn.jsdelivr.net https://ipapi.co;"
    )
    response.headers['Content-Security-Policy'] = csp
    return response

# --- Action Reversal Helpers ---

def perform_undo(action):
    table = action['target_table']
    a_type = action['action_type']
    old_data = json.loads(action['old_data']) if action['old_data'] else None
    target_id = action['target_id']

    try:
        if a_type == 'add':
            commit_db(f"DELETE FROM {table} WHERE id = ?", (target_id,))
        elif a_type == 'edit':
            if old_data:
                cols = ", ".join([f"{k} = ?" for k in old_data.keys()])
                vals = list(old_data.values())
                vals.append(target_id)
                commit_db(f"UPDATE {table} SET {cols} WHERE id = ?", tuple(vals))
        elif a_type == 'delete':
            if table == 'users' and old_data:
                user = old_data.get('user')
                if user:
                    cols = ", ".join(user.keys())
                    placeholders = ", ".join(["?"] * len(user))
                    commit_db(f"INSERT INTO users ({cols}) VALUES ({placeholders})", tuple(user.values()))
                    attendance = old_data.get('attendance', [])
                    for att in attendance:
                        a_cols = ", ".join(att.keys())
                        a_placeholders = ", ".join(["?"] * len(att))
                        commit_db(f"INSERT INTO attendance ({a_cols}) VALUES ({a_placeholders})", tuple(att.values()))

                    courses = old_data.get('courses', [])
                    for c in courses:
                        c_cols = ", ".join(c.keys())
                        c_placeholders = ", ".join(["?"] * len(c))
                        commit_db(f"INSERT INTO lecturer_courses ({c_cols}) VALUES ({c_placeholders})", tuple(c.values()))
            else:
                if old_data:
                    cols = ", ".join(old_data.keys())
                    placeholders = ", ".join(["?"] * len(old_data))
                    commit_db(f"INSERT INTO {table} ({cols}) VALUES ({placeholders})", tuple(old_data.values()))
        elif a_type == 'toggle':
            if table == 'users':
                commit_db("UPDATE users SET active = 1 - active WHERE id = ?", (target_id,))
        elif a_type == 'bulk_toggle':
            if table == 'users':
                role = old_data.get('role')
                old_status = old_data.get('old_status')
                commit_db("UPDATE users SET active = ? WHERE role = ?", (old_status, role))

        return True, "Success"
    except Exception as e:
        app.logger.error(f"Undo failed: {e}")
        return False, str(e)

def perform_redo(action):
    table = action['target_table']
    a_type = action['action_type']
    new_data = json.loads(action['new_data']) if action['new_data'] else None
    target_id = action['target_id']

    try:
        if a_type == 'add':
            if new_data:
                if 'id' not in new_data:
                    new_data['id'] = target_id
                cols = ", ".join(new_data.keys())
                placeholders = ", ".join(["?"] * len(new_data))
                commit_db(f"INSERT INTO {table} ({cols}) VALUES ({placeholders})", tuple(new_data.values()))
        elif a_type == 'edit':
            if new_data:
                cols = ", ".join([f"{k} = ?" for k in new_data.keys()])
                vals = list(new_data.values())
                vals.append(target_id)
                commit_db(f"UPDATE {table} SET {cols} WHERE id = ?", tuple(vals))
        elif a_type == 'delete':
            commit_db(f"DELETE FROM {table} WHERE id = ?", (target_id,))
            if table == 'users':
                commit_db("DELETE FROM attendance WHERE student_id = ?", (target_id,))
                commit_db('DELETE FROM lecturer_courses WHERE lecturer_id = ?', (target_id,))
        elif a_type == 'toggle':
             if table == 'users':
                commit_db("UPDATE users SET active = 1 - active WHERE id = ?", (target_id,))
        elif a_type == 'bulk_toggle':
             if table == 'users':
                role = new_data.get('role')
                new_status = new_data.get('new_status')
                commit_db("UPDATE users SET active = ? WHERE role = ?", (new_status, role))

        return True, "Success"
    except Exception as e:
        app.logger.error(f"Redo failed: {e}")
        return False, str(e)

_SYNC_STATUS = {"running": False, "message": "Idle", "progress": 0, "total": 0}

# --- Routes & Views ---

@app.route('/api/admin/recent-actions')
def api_recent_actions():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify([])

    actions = query_db("""
        SELECT a.*, u.first_name, u.last_name
        FROM action_history a
        LEFT JOIN users u ON a.user_id = u.id
        ORDER BY a.timestamp DESC LIMIT 10
    """)

    out = []
    for a in actions:
        d = dict(a)
        if isinstance(d['timestamp'], datetime):
            d['timestamp'] = d['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
        out.append(d)
    return jsonify(out)

@app.route('/api/admin/security-alerts')
def api_security_alerts():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify([])

    alerts = query_db("""
        SELECT s.*, u.first_name, u.last_name, u.ub_id, u.role as user_role
        FROM security_alerts s
        LEFT JOIN users u ON s.user_id = u.id
        ORDER BY s.created_at DESC LIMIT 50
    """)

    out = []
    for a in alerts:
        d = dict(a)
        if isinstance(d['created_at'], datetime):
            d['created_at'] = d['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        out.append(d)
    return jsonify(out)

@app.route('/api/admin/block-user', methods=['POST'])
def api_block_user():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    user_id = data.get('user_id')
    duration = data.get('duration') # integer
    unit = data.get('unit') # 'hours' or 'days'

    if not user_id or not duration:
        return jsonify({'success': False, 'message': 'Missing data'}), 400

    try:
        duration = int(duration)
        if unit == 'days':
            until = datetime.now(timezone.utc) + timedelta(days=duration)
        else:
            until = datetime.now(timezone.utc) + timedelta(hours=duration)

        commit_db("UPDATE users SET blocked_until = ? WHERE id = ?", (until, user_id))

        # Log the action
        log_action(g.user['id'], 'block', 'users', user_id, None, {'blocked_until': until.isoformat()})

        # Update any related alerts to 'resolved'
        commit_db("UPDATE security_alerts SET status = 'blocked' WHERE user_id = ? AND status = 'open'", (user_id,))

        return jsonify({'success': True, 'message': f'User blocked until {until.strftime("%Y-%m-%d %H:%M:%S")} UTC'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/unblock-user', methods=['POST'])
def api_unblock_user():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    user_id = data.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID required'}), 400

    commit_db("UPDATE users SET blocked_until = NULL WHERE id = ?", (user_id,))
    log_action(g.user['id'], 'unblock', 'users', user_id, None, None)

    return jsonify({'success': True, 'message': 'User unblocked successfully'})

@app.route('/api/admin/analytics/overview')
def api_analytics_overview():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({})
    total_students = query_db("SELECT COUNT(*) as c FROM users WHERE role='student'", one=True)['c']
    total_lecturers = query_db("SELECT COUNT(*) as c FROM users WHERE role='lecturer'", one=True)['c']
    total_sessions = query_db("SELECT COUNT(*) as c FROM sessions", one=True)['c']
    total_attendance = query_db("SELECT COUNT(*) as c FROM attendance", one=True)['c']
    today_attendance = query_db("SELECT COUNT(*) as c FROM attendance WHERE timestamp::date = CURRENT_DATE", one=True)['c']
    active_sessions = query_db("SELECT COUNT(*) as c FROM sessions WHERE active=1", one=True)['c']
    attendance_rate = round((total_attendance / (total_sessions * total_students) * 100), 1) if total_sessions and total_students else 0
    return jsonify({
        'total_students': total_students,
        'total_lecturers': total_lecturers,
        'total_sessions': total_sessions,
        'total_attendance': total_attendance,
        'today_attendance': today_attendance,
        'active_sessions': active_sessions,
        'attendance_rate': attendance_rate
    })

@app.route('/api/admin/analytics/attendance-trend')
def api_analytics_trend():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify([])
    trend = query_db("""
        SELECT DATE(timestamp) as day, COUNT(*) as count
        FROM attendance
        WHERE timestamp >= NOW() - INTERVAL '30 days'
        GROUP BY DATE(timestamp)
        ORDER BY day ASC
    """)
    return jsonify([{'day': str(r['day']), 'count': r['count']} for r in trend])

@app.route('/api/admin/analytics/top-students')
def api_analytics_top_students():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify([])
    top = query_db("""
        SELECT u.id, u.first_name, u.last_name, u.ub_id, COUNT(a.id) as attendance_count, COALESCE(u.reward_points,0) as reward_points
        FROM users u
        LEFT JOIN attendance a ON a.student_id = u.id
        WHERE u.role = 'student'
        GROUP BY u.id
        ORDER BY attendance_count DESC
        LIMIT 10
    """)
    return jsonify([dict(r) for r in top])

@app.route('/admin/undo', methods=['POST'])
def admin_undo():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    action_id = data.get('action_id')

    action = query_db("SELECT * FROM action_history WHERE id = ? AND reverted = 0", (action_id,), one=True)
    if not action:
        return jsonify({'success': False, 'message': 'Action not found or already reverted'}), 404

    success, msg = perform_undo(action)

    if success:
        commit_db("UPDATE action_history SET reverted = 1 WHERE id = ?", (action_id,))
        return jsonify({'success': True, 'message': 'Action undone successfully'})
    else:
        return jsonify({'success': False, 'message': msg})

@app.route('/api/intel/report', methods=['POST'])
@csrf.exempt # Exempt from CSRF as it's an automated telemetry endpoint
def api_intel_report():
    try:
        data = request.get_json() or {}
        user_id = session.get('user_id')
        ip = request.remote_addr or 'unknown'
        ua = request.headers.get('User-Agent', 'unknown')

        # Ensure coordinates are floats or None
        lat = data.get('lat')
        lon = data.get('lon')
        try:
            if lat is not None: lat = float(lat)
            if lon is not None: lon = float(lon)
        except (ValueError, TypeError):
            lat = None
            lon = None

        url = data.get('url')
        device_info = json.dumps(data.get('device_info', {}))

        commit_db("INSERT INTO traffic_intelligence (user_id, ip, ua, lat, lon, url, device_info) VALUES (?,?,?,?,?,?,?)",
                  (user_id, ip, ua, lat, lon, url, device_info))

        # Update users table with latest identity telemetry for the Identity Vault
        ext_ip = data.get('device_info', {}).get('ip', ip)
        if user_id:
            commit_db("UPDATE users SET current_ip = ?, last_activity = ? WHERE id = ?",
                      (ext_ip, datetime.now(timezone.utc), user_id))

        # VPN Detection Engine
        vpn_detected = False
        if ext_ip and ext_ip != '127.0.0.1' and ext_ip != 'unknown':
            try:
                import requests
                # Use ip-api.com for detection (Free for non-commercial, 45 requests/min)
                # Cache results for 10 minutes to avoid hitting limits
                v_resp = requests.get(f"http://ip-api.com/json/{ext_ip}?fields=proxy,hosting,mobile", timeout=5)
                if v_resp.status_code == 200:
                    res = v_resp.json()
                    if res.get('proxy') or res.get('hosting'):
                        vpn_detected = True
                        if user_id:
                            # Avoid duplicate alerts for the same user/IP in the last hour
                            existing = query_db("SELECT id FROM security_alerts WHERE user_id = ? AND ip = ? AND type = 'VPN_DETECTED' AND created_at > NOW() - INTERVAL '1 hour'", (user_id, ext_ip), one=True)
                            if not existing:
                                commit_db("INSERT INTO security_alerts (user_id, type, details, ip) VALUES (?,?,?,?)",
                                          (user_id, 'VPN_DETECTED', f"VPN/Proxy detected on session. ISP: {data.get('device_info', {}).get('isp')}", ext_ip))
            except Exception as v_err:
                app.logger.warning(f"VPN detection failure: {v_err}")

        return jsonify({'success': True, 'vpn_detected': vpn_detected}), 200
    except Exception as e:
        app.logger.error(f"Intel report error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/superadmin/deployment')
def deployment_console():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))
    return render_template('deployment.html')

@app.route('/superadmin/identity-vault')
def identity_vault():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    users = query_db("""
        SELECT id, role, username, first_name, last_name, ub_id, matric, current_ip, last_activity, password_hash
        FROM users
        ORDER BY last_activity DESC NULLS LAST
    """)
    return render_template('identity_vault.html', users=users)

@app.route('/superadmin/vercel/envs', methods=['GET', 'POST'])
def vercel_env_manager():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    t = g.user.get('vercel_token'); pid = g.user.get('vercel_project_id')
    if not t or not pid: return jsonify({'success': False, 'message': 'Vercel config missing'}), 400
    import requests
    h = {"Authorization": f"Bearer {t}", "Content-Type": "application/json"}
    if request.method == 'GET':
        r = requests.get(f"https://api.vercel.com/v8/projects/{pid}/env", headers=h)
        return jsonify({'success': True, 'envs': r.json().get('envs', [])}) if r.status_code == 200 else jsonify({'success': False, 'message': r.text})
    d = request.get_json()
    p = {"key": d.get('key'), "value": d.get('value'), "type": "plain", "target": ["production", "preview", "development"]}
    r = requests.post(f"https://api.vercel.com/v10/projects/{pid}/env", headers=h, json=p)
    return jsonify({'success': True}) if r.status_code in [200, 201] else jsonify({'success': False, 'message': r.text})

@app.route('/superadmin/vercel/envs/<env_id>', methods=['DELETE'])
def vercel_delete_env(env_id):
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    import requests
    h = {"Authorization": f"Bearer {g.user.get('vercel_token')}"}
    r = requests.delete(f"https://api.vercel.com/v9/projects/{g.user.get('vercel_project_id')}/env/{env_id}", headers=h)
    return jsonify({'success': True}) if r.status_code == 200 else jsonify({'success': False, 'message': r.text})

@app.route('/superadmin/intelligence')
def superadmin_intelligence():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    logs = query_db("""
        SELECT t.*, u.first_name, u.last_name, u.role as user_role, u.ub_id
        FROM traffic_intelligence t
        LEFT JOIN users u ON t.user_id = u.id
        ORDER BY t.timestamp DESC LIMIT 500
    """)

    # Parse device info for template
    for log in logs:
        try:
            log['device_details'] = json.loads(log['device_info']) if log['device_info'] else {}
        except Exception:
            log['device_details'] = {}

    return render_template('intelligence.html', logs=logs)

@app.route('/admin/redo', methods=['POST'])
def admin_redo():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    action_id = data.get('action_id')

    action = query_db("SELECT * FROM action_history WHERE id = ? AND reverted = 1", (action_id,), one=True)
    if not action:
        return jsonify({'success': False, 'message': 'Action not found or not reverted'}), 404

    success, msg = perform_redo(action)

    if success:
        commit_db("UPDATE action_history SET reverted = 0 WHERE id = ?", (action_id,))
        return jsonify({'success': True, 'message': 'Action redone successfully'})
    else:
        return jsonify({'success': False, 'message': msg})

@app.route('/health')
def health_check():
    """Endpoint to verify if the application and database connection are ready."""
    try:
        db = get_db()
        with db.cursor() as cur:
            cur.execute('SELECT 1')
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now(timezone.utc).isoformat()
        }), 200
    except Exception as e:
        app.logger.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': datetime.now(timezone.utc).isoformat()
        }), 500


# New Landing Page Route
@app.route('/')
def index():
    return render_template('index.html')

# Portal Hub (Previous index)
@app.route('/portal')
def portal_hub():
    # Fetch team for slideshow
    members = []
    try:
        members = query_db('''
            SELECT m.*, r.name as role_name
            FROM team_members m
            JOIN team_roles r ON m.role_id = r.id
            WHERE m.visible = 1
        ''')
    except Exception:
        pass
    return render_template('portal_hub.html', team_members=members)

# Login for student/lecturer/admin
@app.route('/login/<role>', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
@csrf.exempt
def login(role):
    if role not in ['student', 'lecturer', 'admin', 'superadmin']:
        return "Invalid role", 400

    # If already logged in as this role, redirect to dashboard
    if g.user and g.user['role'] == role:
        if role == 'lecturer': return redirect(url_for('lecturer_dashboard'))
        if role == 'student': return redirect(url_for('student_dashboard'))
        if role == 'admin': return redirect(url_for('admin_dashboard'))
        if role == 'superadmin': return redirect(url_for('superadmin_dashboard'))

    if request.method == 'POST':
        if role == 'student':
            # Defensive parsing to avoid hard 400s when expected fields are missing
            identifier = (request.form.get('identifier') or '').strip()
            pw = request.form.get('password')
            if not identifier or not pw:
                flash('Invalid login request. Please provide your UB ID / Matric and password.')
                return redirect(url_for('login', role=role))

            user = query_db("SELECT * FROM users WHERE (ub_id = ? OR matric = ?) AND role = 'student'", (identifier, identifier), one=True)
            if user and user['active'] and check_password_hash(user['password_hash'], pw):

                security_log("SUCCESSFUL_LOGIN", f"Role: student, ID: {identifier}")

                # Clear failed logins on successful entry
                ip = request.remote_addr or 'unknown'
                commit_db("DELETE FROM failed_logins WHERE ip = ?", (ip,))

                session.clear()
                session['user_id'] = user['id']
                session.permanent = True # Keep students signed in
                ip = request.remote_addr or 'unknown'
                commit_db('UPDATE users SET current_ip = ?, last_activity = ? WHERE id = ?', (ip, datetime.now(timezone.utc), user['id']))
                return redirect(url_for('student_dashboard'))
            # Brute-force blocking toggleable
            ip = request.remote_addr or 'unknown'
            security_mode = get_setting('security_mode', 'off')

            if security_mode == 'on':
                commit_db("INSERT INTO failed_logins (ip, username) VALUES (?,?)", (ip, identifier))
                window = datetime.now(timezone.utc) - timedelta(minutes=10)
                fails_res = query_db("SELECT COUNT(*) as c FROM failed_logins WHERE ip = ? AND attempted_at > ?", (ip, window), one=True)
                fails = fails_res['c'] if fails_res else 0
                if fails >= 5:
                    commit_db("INSERT INTO blocked_ips (ip, reason) VALUES (?,?) ON CONFLICT (ip) DO NOTHING", (ip, "Brute-force attempt detected (UB ID login)"))
                    security_log("IP_BLOCKED", f"Reason: Brute-force, UB ID login: {ub_id}")
                    return "Suspicious activity detected. Your IP has been blocked.", 403

            flash('Invalid UB ID or password')
        else:
            username = request.form['username']
            pw = request.form['password']
            user = query_db('SELECT * FROM users WHERE username = ? AND role = ?', (username, role), one=True)
            if user and user['active'] and check_password_hash(user['password_hash'], pw):
                security_log("SUCCESSFUL_LOGIN", f"Role: {role}, Username: {username}")
                session.clear()
                session['user_id'] = user['id']
                if role == 'lecturer':
                    session.permanent = True # Keep lecturers signed in
                    return redirect(url_for('lecturer_dashboard'))
                elif role == 'superadmin':
                    session.permanent = True
                    return redirect(url_for('superadmin_dashboard'))
                else:
                    # Admins do NOT stay signed in persistently (fulfilling "except admin route")
                    session.permanent = False
                    return redirect(url_for('admin_dashboard'))

            # Brute-force blocking toggleable
            ip = request.remote_addr or 'unknown'
            security_mode = get_setting('security_mode', 'off')

            if security_mode == 'on':
                commit_db("INSERT INTO failed_logins (ip, username) VALUES (?,?)", (ip, username))
                window = datetime.now(timezone.utc) - timedelta(minutes=10)
                fails_res = query_db("SELECT COUNT(*) as c FROM failed_logins WHERE ip = ? AND attempted_at > ?", (ip, window), one=True)
                fails = fails_res['c'] if fails_res else 0
                if fails >= 5:
                    commit_db("INSERT INTO blocked_ips (ip, reason) VALUES (?,?) ON CONFLICT (ip) DO NOTHING", (ip, f"Brute-force attempt detected ({role} login)"))
                    security_log("IP_BLOCKED", f"Reason: Brute-force, {role} login: {username}")
                    return "Suspicious activity detected. Your IP has been blocked.", 403

            flash('Invalid credentials')
    return render_template('login.html', role=role)

@app.route('/logout')
def logout():
    if g.user:
        commit_db('UPDATE users SET current_ip = NULL WHERE id = ?', (g.user['id'],))
    session.clear()
    return redirect(url_for('portal_hub'))

@app.route('/api/team/cycling')
def api_team_cycling():
    # Helper for the footer carousel
    members = get_team_members_with_fallback(is_admin=False)
    out = []
    for m in members:
        out.append({
            'name': m['name'],
            'role': m['role_name']
        })
    return jsonify(out)

@app.route('/api/team/roles')
def api_team_roles():
    try:
        roles = query_db('SELECT * FROM team_roles ORDER BY id ASC')
        return jsonify([dict(r) for r in roles])
    except Exception:
        return jsonify([])

@app.route('/api/team/social-links')
def api_team_social_links():
    # Helper to get all team members for public display
    members = get_team_members_with_fallback(is_admin=False)
    # Filter only necessary public info
    public_members = []
    for m in members:
        public_members.append({
            'name': m['name'],
            'role': m['role_name'],
            'email': m.get('email'),
            'whatsapp': m.get('whatsapp')
        })
    return jsonify({
        'platforms': ['Email', 'WhatsApp', 'LinkedIn', 'GitHub', 'Twitter'],
        'members': public_members
    })

def get_team_members_with_fallback(is_admin=False):
    members = []
    try:
        where_clause = '' if is_admin else 'WHERE m.visible = 1'
        members = query_db(f'''
            SELECT m.*, r.name as role_name
            FROM team_members m
            LEFT JOIN team_roles r ON m.role_id = r.id
            {where_clause}
            ORDER BY m.id ASC
        ''')
    except Exception as e:
        app.logger.error(f"Error fetching team members: {e}")

    # Hard-coded fallback for the Founder if DB is empty, missing him, or connection failed
    if not members or not any(m['name'] == 'Omonire O. Great' for m in members):
        default_img = "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iODAwIiBoZWlnaHQ9IjgwMCIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cmVjdCB3aWR0aD0iMTAwJSIgaGVpZ2h0PSIxMDAlIiBmaWxsPSIjZjNmNGY2Ii8+PGNpcmNsZSBjeD0iNDAwIiBjeT0iMzAwIiByPSIxNTAiIGZpbGw9IiNkMWQ1ZGIiLz48cGF0aCBkPSJNNTAsODAwIEM1MCw2MDAgMjAwLDUwMCA0MDAsNTAwIEM2MDAsNTAwIDc1MCw2MDAgNzUwLDgwMCIgZmlsbD0iI2QxZDVkYiIvPjwvc3ZnPg=="
        founder = {
            'id': 9999,
            'name': 'Omonire O. Great',
            'role_name': 'CEO & Founder',
            'bio': OMONIRE_BIO,
            'image': default_img,
            'visible': 1,
            'role_id': 0 # Dummy ID
        }
        if not members:
            members = [founder]
        else:
            members.insert(0, founder)
    return members

@app.route('/meet-the-team')
def meet_the_team():
    # GET request: Display members (Read Only)
    is_admin = g.user and g.user['role'] in ['admin', 'superadmin']
    members = get_team_members_with_fallback(is_admin)
    roles = []
    if is_admin:
        try:
            roles = query_db('SELECT * FROM team_roles')
        except Exception:
            pass
    return render_template('meet_the_team.html', members=members, roles=roles)

@app.route('/admin/team/manage', methods=['POST'])
def admin_manage_team():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized. Superadmin access required.'}), 403

    action = request.form.get('action')

    if action == 'add_role':
        name = request.form.get('role_name')
        if name:
            commit_db('INSERT INTO team_roles (name) VALUES (?) ON CONFLICT DO NOTHING', (name,))
            role_row = query_db('SELECT id FROM team_roles WHERE name = ?', (name,), one=True)
            if role_row:
                log_action(g.user['id'], 'add', 'team_roles', role_row['id'], None, {'name': name})
            flash('Role added successfully')

    elif action == 'edit_role':
        rid = request.form.get('role_id')
        name = request.form.get('role_name')
        if rid and name:
            old_role = query_db('SELECT * FROM team_roles WHERE id = ?', (rid,), one=True)
            commit_db('UPDATE team_roles SET name = ? WHERE id = ?', (name, rid))
            if old_role:
                log_action(g.user['id'], 'edit', 'team_roles', rid, {'name': old_role['name']}, {'name': name})
            flash('Role updated successfully')

    elif action == 'delete_role':
        rid = request.form.get('role_id')
        if rid:
            old_role = query_db('SELECT * FROM team_roles WHERE id = ?', (rid,), one=True)
            # Unset role_id for team members assigned to this role
            commit_db('UPDATE team_members SET role_id = NULL WHERE role_id = ?', (rid,))
            commit_db('DELETE FROM team_roles WHERE id = ?', (rid,))
            if old_role:
                log_action(g.user['id'], 'delete', 'team_roles', rid, old_role, None)
            flash('Role deleted successfully')

    elif action == 'add_member':
        name = request.form.get('name')
        role_id = request.form.get('role_id')
        try:
            role_id = int(role_id) if role_id else None
        except (ValueError, TypeError):
            role_id = None
        bio = request.form.get('bio')
        email = request.form.get('email')
        whatsapp = request.form.get('whatsapp')
        visible = 1 if request.form.get('visible') else 0
        image_b64 = None

        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                img_data = file.read()
                if img_data:
                    b64_str = base64.b64encode(img_data).decode('utf-8')
                    mimetype = file.mimetype or 'image/jpeg'
                    image_b64 = f"data:{mimetype};base64,{b64_str}"

        commit_db('INSERT INTO team_members (role_id, name, bio, image, visible, email, whatsapp) VALUES (?,?,?,?,?,?,?)',
                  (role_id, name, bio, image_b64, visible, email, whatsapp))
        new_m = query_db('SELECT id FROM team_members WHERE name = ? ORDER BY id DESC LIMIT 1', (name,), one=True)
        if new_m:
            log_action(g.user['id'], 'add', 'team_members', new_m['id'], None, {'role_id': role_id, 'name': name, 'bio': bio, 'visible': visible, 'email': email, 'whatsapp': whatsapp})
        flash('Team member added')

    elif action == 'edit_member':
        mid = request.form.get('member_id')
        old_m = query_db('SELECT * FROM team_members WHERE id = ?', (mid,), one=True)
        name = request.form.get('name')
        role_id = request.form.get('role_id')
        try:
            role_id = int(role_id) if role_id else None
        except (ValueError, TypeError):
            role_id = None
        bio = request.form.get('bio')
        email = request.form.get('email')
        whatsapp = request.form.get('whatsapp')
        visible = 1 if request.form.get('visible') else 0

        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                img_data = file.read()
                if img_data:
                    b64_str = base64.b64encode(img_data).decode('utf-8')
                    mimetype = file.mimetype or 'image/jpeg'
                    image_b64 = f"data:{mimetype};base64,{b64_str}"
                    commit_db('UPDATE team_members SET image = ? WHERE id = ?', (image_b64, mid))

        commit_db('UPDATE team_members SET name=?, role_id=?, bio=?, visible=?, email=?, whatsapp=? WHERE id=?',
                  (name, role_id, bio, visible, email, whatsapp, mid))
        if old_m:
            old_data = {k: v for k, v in old_m.items() if k not in ('id', 'image')}
            new_data = {'name': name, 'role_id': role_id, 'bio': bio, 'visible': visible, 'email': email, 'whatsapp': whatsapp}
            log_action(g.user['id'], 'edit', 'team_members', mid, old_data, new_data)
        flash('Team member updated')

    elif action == 'delete_member':
        mid = request.form.get('member_id')
        old_m = query_db('SELECT * FROM team_members WHERE id = ?', (mid,), one=True)
        if old_m:
            old_data = {k: v for k, v in old_m.items() if k != 'id'}
            log_action(g.user['id'], 'delete', 'team_members', mid, old_data, None)
        commit_db('DELETE FROM team_members WHERE id = ?', (mid,))
        flash('Team member removed')

    return redirect(url_for('superadmin_team'))

@app.route('/admin/team/members/add')
def team_member_add():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))
    roles = query_db('SELECT * FROM team_roles')
    return render_template('team_member_form.html', member=None, roles=roles)

@app.route('/admin/team/members/edit/<int:mid>')
def team_member_edit(mid):
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))
    member = query_db('SELECT * FROM team_members WHERE id = ?', (mid,), one=True)
    if not member:
        flash('Team member not found')
        return redirect(url_for('superadmin_team'))
    roles = query_db('SELECT * FROM team_roles')
    return render_template('team_member_form.html', member=member, roles=roles)

@app.route('/admin/team/roles')
def team_roles():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))
    roles = query_db('SELECT * FROM team_roles')
    return render_template('team_role_form.html', roles=roles)

@app.route('/admin/team/roles/edit/<int:rid>')
def team_role_edit(rid):
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))
    role = query_db('SELECT * FROM team_roles WHERE id = ?', (rid,), one=True)
    if not role:
        flash('Role not found')
        return redirect(url_for('team_roles'))
    roles = query_db('SELECT * FROM team_roles')
    return render_template('team_role_form.html', edit_role=role, roles=roles)

@app.route('/meet-the-team/<role>/<name>')
def team_member_profile(role, name):
    # Unquote URL parameters to handle encoded characters like %20 or %26
    role = unquote(role)
    name = unquote(name)

    member = None
    try:
        member = query_db('''
            SELECT m.*, r.name as role_name
            FROM team_members m
            JOIN team_roles r ON m.role_id = r.id
            WHERE r.name = ? AND m.name = ? AND m.visible = 1
        ''', (role, name), one=True)
    except Exception:
        pass

    if not member and name == 'Omonire O. Great':
        default_img = "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iODAwIiBoZWlnaHQ9IjgwMCIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cmVjdCB3aWR0aD0iMTAwJSIgaGVpZ2h0PSIxMDAlIiBmaWxsPSIjZjNmNGY2Ii8+PGNpcmNsZSBjeD0iNDAwIiBjeT0iMzAwIiByPSIxNTAiIGZpbGw9IiNkMWQ1ZGIiLz48cGF0aCBkPSJNNTAsODAwIEM1MCw2MDAgMjAwLDUwMCA0MDAsNTAwIEM2MDAsNTAwIDc1MCw2MDAgNzUwLDgwMCIgZmlsbD0iI2QxZDVkYiIvPjwvc3ZnPg=="
        member = {
            'id': 9999,
            'name': 'Omonire O. Great',
            'role_name': 'CEO & Founder',
            'bio': OMONIRE_BIO,
            'image': default_img,
            'visible': 1,
                'email': FOUNDER_EMAIL,
                'whatsapp': FOUNDER_WHATSAPP
        }

    if not member:
        flash('Team member not found')
        return redirect(url_for('meet_the_team'))
    return render_template('team_member.html', member=member)

@app.route('/owner')
def owner_info():
    return redirect(url_for('meet_the_team'))

@app.route('/forgot-password/<role>', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def forgot_password(role):
    if request.method == 'POST':
        if role == 'student':
            identifier = request.form.get('identifier')
            user = query_db("SELECT * FROM users WHERE (ub_id = ? OR matric = ?) AND role = 'student'", (identifier, identifier), one=True)
        else:
            username = request.form.get('username')
            user = query_db("SELECT * FROM users WHERE username = ? AND role = ?", (username, role), one=True)

        if not user:
            flash('Account not found')
            return redirect(url_for('forgot_password', role=role))

        if not user.get('security_q1') or not user.get('security_q2'):
            flash('This account does not have security questions set. Contact admin for reset.')
            return redirect(url_for('login', role=role))

        session['reset_user_id'] = user['id']
        return redirect(url_for('reset_password'))

    return render_template('forgot_password.html', role=role)

@app.route('/reset-password', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def reset_password():
    user_id = session.get('reset_user_id')
    if not user_id:
        return redirect(url_for('portal_hub'))

    user = query_db("SELECT * FROM users WHERE id = ?", (user_id,), one=True)
    if not user:
        return redirect(url_for('portal_hub'))

    if request.method == 'POST':
        a1 = request.form.get('security_a1', '').strip().lower()
        a2 = request.form.get('security_a2', '').strip().lower()

        if a1 == user['security_a1'] and a2 == user['security_a2']:
            # Success
            session['identity_verified'] = True
            return render_template('reset_password.html', stage='new_password', reset_token=secrets.token_hex(16))
        else:
            flash('Incorrect answers to security questions.')
            return redirect(url_for('reset_password'))

    return render_template('reset_password.html', stage='verify',
                           security_q1=user['security_q1'],
                           security_q2=user['security_q2'],
                           user_id=user_id)

@app.route('/set-new-password', methods=['POST'])
def set_new_password():
    if not session.get('identity_verified') or not session.get('reset_user_id'):
        return redirect(url_for('portal_hub'))

    pw = request.form.get('password')
    if not pw:
        flash('Password is required')
        return redirect(url_for('reset_password'))

    user_id = session.get('reset_user_id')
    commit_db("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(pw), user_id))

    session.pop('reset_user_id', None)
    session.pop('identity_verified', None)

    flash('Password updated successfully. Please login.')
    # Redirect to appropriate login based on user role (need to fetch it first or store it in session)
    user = query_db("SELECT role FROM users WHERE id = ?", (user_id,), one=True)
    role = user['role'] if user else 'student'
    return redirect(url_for('login', role=role))

# Signup for student / lecturer
@app.route('/signup/<role>', methods=['GET','POST'])
@limiter.limit("5 per hour")
def signup(role):
    if role not in ('student','lecturer'):
        return 'Invalid role', 400
    if request.method == 'POST':
        first = request.form.get('first_name')
        middle = request.form.get('middle_name')
        last = request.form.get('last_name')
        pw = request.form.get('password')

        sq1 = request.form.get('security_q1')
        sa1 = request.form.get('security_a1')
        sq2 = request.form.get('security_q2')
        sa2 = request.form.get('security_a2')

        if not first or not last or not pw or not sq1 or not sa1 or not sq2 or not sa2:
            flash('Missing required fields, including security questions')
            return redirect(url_for('signup', role=role))

        # Normalize answers for easier recovery
        sa1 = sa1.strip().lower()
        sa2 = sa2.strip().lower()

        if role == 'student':
            ub = request.form.get('ub_id')
            if not ub:
                flash('UB ID required')
                return redirect(url_for('signup', role=role))
            # ensure UB unique
            exists = query_db('SELECT * FROM users WHERE ub_id = ?', (ub,), one=True)
            if exists:
                flash('UB ID already exists')
                return redirect(url_for('signup', role=role))
            # optional department selection
            dept_id = request.form.get('department_id')
            try:
                dept_val = int(dept_id) if dept_id else None
            except Exception:
                dept_val = None
            matric = request.form.get('matric') or None
            commit_db('INSERT INTO users (role, username, password_hash, first_name, middle_name, last_name, ub_id, department_id, matric, security_q1, security_a1, security_q2, security_a2) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                      ('student', None, generate_password_hash(pw), first, middle, last, ub, dept_val, matric, sq1, sa1, sq2, sa2))
            flash('Student account created. Please login')
            return redirect(url_for('login', role='student'))
        else:
            username = request.form.get('username') or f"{first.lower()}.{last.lower()}"
            # ensure username unique
            exists = query_db('SELECT * FROM users WHERE username = ?', (username,), one=True)
            if exists:
                flash('Username already exists, choose another')
                return redirect(url_for('signup', role=role))
            db = get_db()
            with db.cursor() as cur:
                cur.execute('INSERT INTO users (role, username, password_hash, first_name, middle_name, last_name, security_q1, security_a1, security_q2, security_a2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id',
                          ('lecturer', username, generate_password_hash(pw), first, middle, last, sq1, sa1, sq2, sa2))
                lecturer_id = cur.fetchone()[0]
                db.commit()

            # associate courses
            course_ids = request.form.getlist('courses')
            for cid in course_ids:
                commit_db('INSERT INTO lecturer_courses (lecturer_id, course_id) VALUES (?,?)', (lecturer_id, int(cid)))
            flash('Lecturer account created. Please login')
            return redirect(url_for('login', role='lecturer'))
    # GET: show form
    courses = query_db('SELECT * FROM courses') if role == 'lecturer' else []
    faculties = []
    if role == 'student':
        faculties = query_db('SELECT * FROM faculties')
    return render_template('signup.html', role=role, courses=courses, faculties=faculties)

# Lecturer dashboard
@app.route('/lecturer')
def lecturer_dashboard():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    lecturer_id = g.user['id']
    sessions = query_db('SELECT * FROM sessions WHERE lecturer_id = ? ORDER BY created_at DESC', (lecturer_id,))
    locations = query_db('SELECT * FROM locations WHERE lecturer_id = ?', (lecturer_id,))
    courses = query_db('SELECT * FROM courses')
    return render_template('lecturer.html', locations=locations, sessions=sessions, courses=courses)

@app.route('/session/create', methods=['POST'])
def create_session():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    token = generate_token(8)
    # Course: prefer selected course_id if provided, otherwise use typed course_name
    course_id = request.form.get('course_id')
    if course_id and course_id != 'other':
        try:
            cid = int(course_id)
            course_row = query_db('SELECT * FROM courses WHERE id = ?', (cid,), one=True)
            course_name = course_row['name'] if course_row else request.form.get('course_name')
            course_id_val = cid
        except Exception:
            course_name = request.form.get('course_name')
            course_id_val = None
    else:
        course_name = request.form.get('course_name') or 'Untitled Course'
        course_id_val = None

    host = request.form.get('host') or ''
    duration = int(request.form.get('duration') or DEFAULT_SESSION_DURATION_MIN)
    grace = int(request.form.get('grace') or DEFAULT_GRACE_MIN)
    mode = request.form.get('attendance_mode') or 'both'
    strict = int(request.form.get('strict_mode') or 1)

    # Coordinates are optional for 'token' mode
    lat_raw = request.form.get('latitude')
    lon_raw = request.form.get('longitude')
    lat = float(lat_raw) if lat_raw else 0.0
    lon = float(lon_raw) if lon_raw else 0.0

    radius = float(request.form.get('radius') or DEFAULT_RADIUS_M)
    start_time = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    commit_db('INSERT INTO sessions (token, lecturer_id, course_id, course_name, host, start_time, duration_min, grace_min, latitude, longitude, radius_m, attendance_mode, strict_mode) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
              (token, g.user['id'], course_id_val, course_name, host, start_time, duration, grace, lat, lon, radius, mode, strict))
    # Notify all students with registered emails about the new session
    try:
        students = query_db("SELECT email, first_name, last_name FROM users WHERE role='student' AND email IS NOT NULL AND email != ''")
        lecturer_name = f"{g.user.get('first_name','')} {g.user.get('last_name','')}".strip() or 'Your lecturer'
        for st in students:
            send_session_notification(st['email'], f"{st['first_name']} {st['last_name']}", course_name, token, lecturer_name)
    except Exception as e:
        app.logger.error(f"Session notification error: {e}")
    flash(f"Session created with token: {token}")
    return redirect(url_for('lecturer_dashboard'))

# Geocode endpoint using OpenStreetMap Nominatim (no API key required). Respects rate limits.
@app.route('/geocode')
def geocode_endpoint():
    q = request.args.get('q')
    if not q:
        return jsonify({'success': False, 'message': 'Query required'}), 400
    # Use a slightly longer timeout to accommodate slow responses; return friendly messages on failure
    try:
        res = geocode(q, timeout=15)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Geocode error: {e}. Try entering coordinates manually.'}), 500
    if not res.get('success'):
        msg = res.get('message','Geocode failed')
        # if it's a requests timeout or network issue, give actionable advice
        if 'timed out' in msg.lower() or 'timeout' in msg.lower():
            return jsonify({'success': False, 'message': 'Geocoding timed out. Try again or enter coordinates manually.'}), 504
        return jsonify({'success': False, 'message': msg}), 404
    return jsonify(res), 200

# View a session (lecturer) with live log and simple chart
@app.route('/session/<int:session_id>')
def session_detail(session_id):
    s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)
    if not s:
        return 'Session not found', 404
    s = dict(s)
    # Convert datetime objects to strings for template/JS compatibility
    if isinstance(s.get('start_time'), datetime):
        s['start_time'] = s['start_time'].strftime('%Y-%m-%d %H:%M:%S')

    # If course_id present, fetch course label
    if 'course_id' in s and s['course_id']:
        c = query_db('SELECT * FROM courses WHERE id = ?', (s['course_id'],), one=True)
        if c:
            s['course_name'] = c['name']
    # ensure host is available
    s['host'] = s['host'] if 'host' in s.keys() else ''
    s['active'] = s['active'] if 'active' in s.keys() else 1
    # attendance counts for charts (simple weekly counts based on timestamp)
    rows = query_db('SELECT timestamp FROM attendance WHERE session_id = ?', (session_id,))
    timestamps = [parse_dt(r['timestamp']) for r in rows] if rows else []
    return render_template('session.html', s=s, session_id=session_id)

@app.route('/session/<int:session_id>/terminate', methods=['POST'])
def terminate_session(session_id):
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)
    if not s:
        flash('Session not found')
        return redirect(url_for('lecturer_dashboard'))
    if s['lecturer_id'] != g.user['id']:
        flash('You are not allowed to terminate this session')
        return redirect(url_for('lecturer_dashboard'))
    try:
        commit_db('UPDATE sessions SET active = 0 WHERE id = ?', (session_id,))
        # audit
        try:
            app.logger.info(f"ADMIN_AUDIT: {datetime.now(timezone.utc).isoformat()} - lecturer {g.user['id']} ({g.user.get('first_name')}) terminated session {session_id}")
        except Exception:
            pass
        flash('Session has been terminated')
    except Exception as e:
        flash('Failed to terminate session: ' + str(e))
    return redirect(url_for('lecturer_dashboard'))

@app.route('/session/<int:session_id>/delete', methods=['POST'])
def delete_session_route(session_id):
    if not g.user or g.user['role'] not in ['lecturer', 'admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)
    if not s:
        return jsonify({'success': False, 'message': 'Session not found'}), 404

    # Permission check: lecturer must own it, others can delete any
    if g.user['role'] == 'lecturer' and s['lecturer_id'] != g.user['id']:
        return jsonify({'success': False, 'message': 'You are not allowed to delete this session'}), 403

    try:
        # Fetch related data for undo logging
        attendance = query_db("SELECT * FROM attendance WHERE session_id = ?", (session_id,))
        # Convert objects to strings for JSON
        for att in attendance:
            for k, v in att.items():
                if isinstance(v, datetime):
                    att[k] = v.strftime('%Y-%m-%d %H:%M:%S')

        # Log the action
        log_action(g.user['id'], 'delete', 'sessions', session_id, {'session': s, 'attendance': attendance}, None)

        # Delete attendance first
        commit_db('DELETE FROM attendance WHERE session_id = ?', (session_id,))
        # Delete session
        commit_db('DELETE FROM sessions WHERE id = ?', (session_id,))

        # Security log
        security_log("SESSION_DELETED", f"Session {session_id} ({s['course_name']}) deleted by {g.user['role']} {g.user['id']}")

        return jsonify({'success': True, 'message': 'Session and associated attendance records deleted permanently.'})
    except Exception as e:
        app.logger.error(f"Error deleting session {session_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/session/<int:session_id>/stats')
def session_stats(session_id):
    # Return simple counts (for last 7 entries) for bar chart
    rows = query_db('SELECT timestamp FROM attendance WHERE session_id = ? ORDER BY timestamp ASC', (session_id,))
    labels = []
    counts = []
    for i, r in enumerate(rows):
        labels.append(str(i+1))
        counts.append(1)
    if not labels:
        labels = ['No data']
        counts = [0]
    return jsonify({'labels': labels, 'counts': counts})

@app.route('/session/<int:session_id>/live')
def session_live(session_id):
    rows = query_db('SELECT first_name, middle_name, last_name, timestamp FROM attendance WHERE session_id = ? ORDER BY timestamp DESC LIMIT 10', (session_id,))
    names = []
    for r in rows:
        n = ' '.join(filter(None, [r['first_name'], r['middle_name'], r['last_name']]))
        names.append(f"{n} - {r['timestamp']}")
    return jsonify({'names': names})

@app.route('/session/<int:session_id>/mark-manual', methods=['POST'])
def mark_attendance_manual(session_id):
    if not g.user or g.user['role'] not in ['lecturer', 'admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    identifier = data.get('identifier', '').strip()
    if not identifier:
        return jsonify({'success': False, 'message': 'UB ID or Matric Number required'}), 400

    s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)
    if not s:
        return jsonify({'success': False, 'message': 'Session not found'}), 404

    # Check if user exists
    user = query_db("SELECT * FROM users WHERE role = 'student' AND (ub_id = ? OR matric = ?)", (identifier, identifier), one=True)
    if not user:
        return jsonify({'success': False, 'message': 'Student not found with provided ID'}), 404

    # Check if already marked
    already_marked = query_db('SELECT * FROM attendance WHERE session_id = ? AND student_id = ?', (session_id, user['id']), one=True)
    if already_marked:
        return jsonify({'success': False, 'message': 'Student already marked present for this session'}), 409

    # Mark as present (Manual)
    timestamp = datetime.now(timezone.utc).isoformat()
    commit_db('INSERT INTO attendance (session_id, student_id, timestamp, ip, device_fp, lat, lon, status, first_name, middle_name, last_name) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
              (session_id, user['id'], timestamp, '0.0.0.0', 'manual-override', s['latitude'], s['longitude'], 'manual', user['first_name'], user['middle_name'], user['last_name']))

    # Audit log
    try:
        app.logger.info(f"ADMIN_AUDIT: {datetime.now(timezone.utc).isoformat()} - lecturer {g.user['id']} manually marked student {user['id']} for session {session_id}")
    except Exception:
        pass

    return jsonify({'success': True, 'message': f"Student {user['first_name']} {user['last_name']} marked present."})

# Export attendance report (Excel or PDF)
@app.route('/session/<int:session_id>/export')
def export_session(session_id):
    t = request.args.get('t','excel')
    # Join with users, departments, and faculties to get extra info
    rows = query_db('''
        SELECT a.*, d.name as department_name, f.name as faculty_name
        FROM attendance a
        JOIN users u ON a.student_id = u.id
        LEFT JOIN departments d ON u.department_id = d.id
        LEFT JOIN faculties f ON d.faculty_id = f.id
        WHERE a.session_id = ?
        ORDER BY a.timestamp ASC
    ''', (session_id,))
    data = []
    for idx, r in enumerate(rows, start=1):
        data.append({
            'S/N': idx,
            'First Name': r['first_name'],
            'Middle Name': r['middle_name'] or '',
            'Last Name': r['last_name'],
            'Faculty': r['faculty_name'] or 'N/A',
            'Department': r['department_name'] or 'N/A',
            'Timestamp': r['timestamp']
        })
    if t == 'excel':
        if pd:
            df = pd.DataFrame(data)
            bio = BytesIO()
            df.to_excel(bio, index=False, engine='openpyxl')
            bio.seek(0)
            return send_file(bio, as_attachment=True, download_name=f'session_{session_id}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        elif Workbook is not None:
            wb = Workbook()
            ws = wb.active
            ws.append(['S/N','First Name','Middle Name','Last Name','Faculty','Department','Timestamp'])
            for row in data:
                ws.append([row['S/N'], row['First Name'], row['Middle Name'], row['Last Name'], row['Faculty'], row['Department'], row['Timestamp']])
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(bio, as_attachment=True, download_name=f'session_{session_id}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            # fallback to CSV
            csv = 'S/N,First Name,Middle Name,Last Name,Faculty,Department,Timestamp\n'
            for row in data:
                csv += f"{row['S/N']},{row['First Name']},{row['Middle Name']},{row['Last Name']},{row['Faculty']},{row['Department']},{row['Timestamp']}\n"
            return send_file(BytesIO(csv.encode()), as_attachment=True, download_name=f'session_{session_id}.csv', mimetype='text/csv')
    elif t == 'pdf':
        if FPDF:
            pdf = FPDF(orientation='L') # Landscape to fit more columns
            pdf.add_page()
            pdf.set_font('Arial','B',12)
            pdf.cell(0,10,f"Attendance Report - Session {session_id}",ln=True)
            pdf.set_font('Arial','B',10)
            # Headers
            pdf.cell(15,8,"S/N",1)
            pdf.cell(30,8,"First Name",1)
            pdf.cell(30,8,"Middle Name",1)
            pdf.cell(30,8,"Last Name",1)
            pdf.cell(55,8,"Faculty",1)
            pdf.cell(55,8,"Department",1)
            pdf.cell(50,8,"Timestamp",1,ln=True)

            pdf.set_font('Arial','',8)
            for r in data:
                pdf.cell(15,8,str(r['S/N']),1)
                pdf.cell(30,8,str(r['First Name']),1)
                pdf.cell(30,8,str(r['Middle Name']),1)
                pdf.cell(30,8,str(r['Last Name']),1)
                pdf.cell(55,8,str(r['Faculty'])[:35],1) # Truncate if too long
                pdf.cell(55,8,str(r['Department'])[:35],1)
                pdf.cell(50,8,str(r['Timestamp']),1,ln=True)
            bio = BytesIO()
            pdf.output(bio)
            bio.seek(0)
            return send_file(bio, as_attachment=True, download_name=f'session_{session_id}.pdf', mimetype='application/pdf')
        else:
            return 'PDF export requires "fpdf" python package. Install it or export Excel/CSV instead.'
    return 'Unsupported export type', 400

@app.route('/api/wallet/assets')
def api_wallet_assets():
    if not g.user:
        return jsonify({'tokens': [], 'nfts': []})
    wallet = g.user.get('wallet_address')
    if not wallet:
        return jsonify({'tokens': [], 'nfts': []})
    return jsonify(get_assets(wallet))

# Student dashboard - shows current sessions and ability to mark attendance
@app.route('/student')
def student_dashboard():
    if not g.user or g.user['role'] != 'student':
        return redirect(url_for('login', role='student'))
    # Find current active session(s)
    now = datetime.now(timezone.utc)
    # Limit to most recent 10 sessions for performance with large user base
    sessions = query_db('SELECT * FROM sessions ORDER BY created_at DESC LIMIT 10')
    # no auto-detect; students will choose which session to join
    s = None
    # Student attendance summary
    total_expected = query_db('SELECT COUNT(*) as c FROM sessions', (), one=True)['c']
    attended = query_db('SELECT COUNT(*) as c FROM attendance WHERE student_id = ?', (g.user['id'],), one=True)['c']
    # Reward data
    reward_points = g.user.get('reward_points') or 0
    streak = g.user.get('streak') or 0
    rank_row = query_db("SELECT COUNT(*) as r FROM users WHERE role = 'student' AND COALESCE(reward_points,0) > ?", (reward_points,), one=True)
    rank = (rank_row['r'] + 1) if rank_row else 0
    return render_template('student.html', s=s, sessions=sessions, total_expected=total_expected, attended=attended, reward_points=reward_points, streak=streak, rank=rank)

@app.route('/leaderboard')
def leaderboard():
    top = query_db("SELECT id, first_name, last_name, ub_id, COALESCE(reward_points,0) as reward_points, COALESCE(streak,0) as streak FROM users WHERE role = 'student' ORDER BY reward_points DESC NULLS LAST LIMIT 50")
    return render_template('leaderboard.html', top=top)

@app.route('/student/upload-card', methods=['POST'])
def upload_card():
    if not g.user or g.user['role'] != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    if 'card_file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400

    file = request.files['card_file']
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400

    # Validate file type
    allowed_extensions = {'png', 'jpg', 'jpeg', 'pdf'}
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in allowed_extensions:
        return jsonify({'success': False, 'message': 'Invalid file format. Use PNG, JPG, or PDF.'}), 400

    try:
        img_data = file.read()
        if not img_data:
            return jsonify({'success': False, 'message': 'File is empty'}), 400

        b64_str = base64.b64encode(img_data).decode('utf-8')
        mimetype = file.mimetype or f"image/{ext}" if ext != 'pdf' else 'application/pdf'
        card_data = f"data:{mimetype};base64,{b64_str}"

        commit_db("INSERT INTO student_cards (student_id, card_data, file_type) VALUES (?,?,?) ON CONFLICT (student_id) DO UPDATE SET card_data = EXCLUDED.card_data, file_type = EXCLUDED.file_type, created_at = NOW()",
                  (g.user['id'], card_data, mimetype))

        return jsonify({'success': True, 'message': 'ID Card uploaded successfully'})
    except Exception as e:
        app.logger.error(f"Card upload error: {e}")
        return jsonify({'success': False, 'message': 'Internal server error during upload'}), 500

@app.route('/student/link-wallet', methods=['POST'])
def link_wallet():
    if not g.user or g.user['role'] != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    data = request.get_json()
    wallet = data.get('wallet_address', '').strip()
    if not wallet:
        commit_db("UPDATE users SET wallet_address = NULL WHERE id = ?", (g.user['id'],))
        return jsonify({'success': True, 'message': 'Wallet unlinked successfully', 'wallet_address': None})
    if not wallet.startswith('5') or len(wallet) < 32:
        return jsonify({'success': False, 'message': 'Invalid wallet address'}), 400
    # Check if wallet already linked to another student
    existing = query_db("SELECT id FROM users WHERE wallet_address = ? AND id != ?", (wallet, g.user['id']), one=True)
    if existing:
        return jsonify({'success': False, 'message': 'This wallet is already linked to another account'}), 409
    commit_db("UPDATE users SET wallet_address = ? WHERE id = ?", (wallet, g.user['id']))
    return jsonify({'success': True, 'message': 'Wallet linked successfully', 'wallet_address': wallet})

@app.route('/student/history')
def student_history():
    if not g.user or g.user['role'] != 'student':
        return redirect(url_for('login', role='student'))

    history = query_db('''
        SELECT a.timestamp, a.status, s.course_name, s.start_time, s.host
        FROM attendance a
        JOIN sessions s ON a.session_id = s.id
        WHERE a.student_id = ?
        ORDER BY a.timestamp DESC
    ''', (g.user['id'],))

    return render_template('student_history.html', history=history)

@app.route('/student/update-email', methods=['POST'])
def update_email():
    if not g.user or g.user['role'] != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    email = request.get_json().get('email', '').strip()
    if email and ('@' not in email or '.' not in email):
        return jsonify({'success': False, 'message': 'Invalid email address'}), 400
    commit_db("UPDATE users SET email = ? WHERE id = ?", (email or None, g.user['id']))
    return jsonify({'success': True, 'message': 'Email saved'})

# Mark attendance endpoint (POST: JSON)
@app.route('/session/mark', methods=['POST'])
@limiter.limit("5 per minute")
def mark_attendance():
    if not g.user or g.user['role'] != 'student':
        return jsonify({'success': False, 'message': 'Not authenticated as student'}), 401
    data = request.get_json()
    token = data.get('token')
    session_id = data.get('session_id')
    lat = data.get('lat')
    lon = data.get('lon')

    s = None
    if session_id:
        s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)

    # Fallback to token lookup if session_id wasn't provided or didn't match
    if not s and token:
        s = query_db('SELECT * FROM sessions WHERE UPPER(token) = UPPER(?)', (token.strip(),), one=True)

    if not s:
        return jsonify({'success': False, 'message': 'Session not found or invalid token'})

    # Mode-based validation
    mode = s.get('attendance_mode', 'both')

    # User-provided token cleanup
    token_str = token.strip().upper() if token else None
    actual_token = s['token'].upper() if s['token'] else ""

    if mode == 'token' or mode == 'both':
        if not token_str or token_str != actual_token:
            return jsonify({'success': False, 'message': 'Invalid or missing session token'})
    if (s['active'] if 'active' in s.keys() else 1) == 0:
        return jsonify({'success': False, 'message': 'This session has been terminated by the lecturer.'})

    # validate location presence only if required by mode
    if s.get('attendance_mode') != 'token':
        if lat is None or lon is None:
            # We don't fail here yet if it's token mode, but 'location' or 'both' MUST have it.
            # Actually, the check later in evaluate_session_for_student handles within_radius.
            # But the user complained about "server evaluation failed" if location was used.
            pass
        try:
            if lat is not None: lat = float(lat)
            if lon is not None: lon = float(lon)
        except Exception:
            return jsonify({'success': False, 'message': 'Invalid location coordinates provided.'}), 400
    else:
        # For token-only, we don't strictly need student coords
        lat = lat if lat is not None else 0.0
        lon = lon if lon is not None else 0.0

    # compute via helper
    try:
        device_id = data.get('device_id')
        info = evaluate_session_for_student(s, g.user['id'], lat, lon, request.remote_addr or 'unknown', request.headers.get('User-Agent',''), device_id)
    except Exception as e:
        app.logger.exception('Error evaluating session for student')
        return jsonify({'success': False, 'message': 'Server error during evaluation. Please try again.'}), 500

    # use info for decisioning below

    # device checks
    ip = request.remote_addr or 'unknown'
    ua = request.headers.get('User-Agent','')
    fp = device_fp(ip, ua)

    # Check if student's UB ID active
    if not g.user['active']:
        return jsonify({'success': False, 'message': 'UB ID not active. Contact admin.'})

    # 'info' was computed earlier using evaluate_session_for_student and is reused here
    # ensure we have a distance_m value available for payload even if None
    info['distance_m'] = info.get('distance_m')

    if info['already_marked']:
        return jsonify({'success': False, 'message': 'You have already marked attendance for this session.'})
    if info['device_conflict']:
        return jsonify({'success': False, 'message': 'This device has already been used for another student for this session.'})
    if info['other_device_used']:
        return jsonify({'success': False, 'message': 'Your UB ID was used from a different device for this session.'})
    if not info['can_mark']:
        return jsonify({'success': False, 'message': info['message']})

    # allowed to mark
    status = 'marked' if info['status'] == 'active' else 'grace_marked'
    # Show approved message for successful marking
    message = 'Attendance Approved' if status == 'marked' else 'Attendance Approved (Grace)'
    grace_seconds = info.get('grace_seconds')

    timestamp = datetime.now(timezone.utc).isoformat()
    fp = device_fp(ip, ua, data.get('device_id'))
    commit_db('INSERT INTO attendance (session_id, student_id, timestamp, ip, device_fp, lat, lon, status, first_name, middle_name, last_name) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
              (s['id'], g.user['id'], timestamp, ip, fp, lat, lon, status, g.user['first_name'], g.user['middle_name'], g.user['last_name']))

    # Reward system: award +5 pts + streak bonuses
    today = datetime.now(timezone.utc).date()
    last_date = g.user.get('last_attendance_date')
    current_streak = g.user.get('streak') or 0
    new_streak = 1
    if last_date:
        last = last_date
        if isinstance(last_date, str):
            from datetime import date as dt_date
            last = dt_date.fromisoformat(last_date)
        diff = (today - last).days
        if diff == 1:
            new_streak = current_streak + 1
        elif diff == 0:
            new_streak = current_streak
        else:
            new_streak = 1

    streak_bonus = 0
    if new_streak >= 10:
        streak_bonus = 10
    elif new_streak >= 5:
        streak_bonus = 5
    elif new_streak >= 3:
        streak_bonus = 2

    points_earned = 5 + streak_bonus
    commit_db('UPDATE users SET reward_points = COALESCE(reward_points,0) + ?, streak = ?, last_attendance_date = ? WHERE id = ?',
              (points_earned, new_streak, today.isoformat(), g.user['id']))

    # Send email receipt
    try:
        send_attendance_receipt(g.user.get('email'), f"{g.user['first_name']} {g.user['last_name']}", s.get('course_name', 'Class'), timestamp, s['id'])
    except Exception:
        pass

    # Solana on-chain integration
    wallet = g.user.get('wallet_address')
    if wallet:
        try:
            mint_token(wallet, 10)
            record_attendance(g.user['id'], s['id'], timestamp, wallet)
            total_att = query_db("SELECT COUNT(*) as c FROM attendance WHERE student_id = ?", (g.user['id'],), one=True)['c']
            milestone = None
            if total_att in [10, 25, 50, 100]:
                tiers = {10: 'bronze', 25: 'silver', 50: 'gold', 100: 'diamond'}
                mint_badge(wallet, tiers[total_att])
                milestone = tiers[total_att]
        except Exception:
            app.logger.exception("Solana integration error")

    payload = {'success': True, 'message': message, 'status': status, 'grace_seconds': grace_seconds, 'distance_m': info.get('distance_m'), 'points_earned': points_earned, 'streak': new_streak}
    return jsonify(payload)

# API: Check session status for a student (helps UI show what will happen before marking)
@app.route('/session/check', methods=['POST'])
def session_check():
    if not g.user or g.user['role'] != 'student':
        return jsonify({'success': False, 'message': 'Not authenticated as student'}), 401
    data = request.get_json() or {}
    token = data.get('token')
    session_id = data.get('session_id')
    lat = data.get('lat')
    lon = data.get('lon')

    if session_id:
        s = query_db('SELECT * FROM sessions WHERE id = ?', (session_id,), one=True)
    else:
        # Case-insensitive token lookup
        s = query_db('SELECT * FROM sessions WHERE UPPER(token) = UPPER(?)', (token.strip() if token else '',), one=True)

    if not s:
        return jsonify({'success': False, 'message': 'Session not found'})
    if (s['active'] if 'active' in s.keys() else 1) == 0:
        return jsonify({'success': False, 'message': 'This session was terminated by the lecturer.'}), 410
    ip = request.remote_addr or 'unknown'
    ua = request.headers.get('User-Agent','')
    device_id = data.get('device_id')
    info = evaluate_session_for_student(s, g.user['id'], lat, lon, ip, ua, device_id)
    info['success'] = True
    return jsonify(info)

# Super admin dashboard
@app.route('/admin')
def admin_dashboard():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))

    # Explicitly fetch for template consistency
    website_status = get_setting('website_status', 'on')
    security_mode = get_setting('security_mode', 'off')

    # Pagination for students to handle 48,000+ users
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    if page < 1:
        page = 1
    per_page = 50
    offset = (page - 1) * per_page

    q = request.args.get('q', '').strip()
    faculty_id = request.args.get('faculty_id')
    department_id = request.args.get('department_id')

    query_base = "FROM users WHERE role = 'student'"
    params = []

    if q:
        query_base += " AND (first_name LIKE ? OR last_name LIKE ? OR ub_id LIKE ? OR matric LIKE ?)"
        params.extend(['%'+q+'%', '%'+q+'%', '%'+q+'%', '%'+q+'%'])

    if faculty_id:
        query_base += " AND (faculty_id = ? OR department_id IN (SELECT id FROM departments WHERE faculty_id = ?))"
        params.extend([faculty_id, faculty_id])

    if department_id:
        query_base += " AND department_id = ?"
        params.append(department_id)

    users = query_db(f"SELECT * {query_base} ORDER BY id DESC LIMIT ? OFFSET ?", tuple(params + [per_page, offset]))
    total_res = query_db(f"SELECT COUNT(*) as c {query_base}", tuple(params), one=True)
    total_users = total_res['c'] if total_res else 0

    total_pages = (total_users + per_page - 1) // per_page

    # Fetch lecturers with search
    l_q = request.args.get('l_q', '').strip()
    if l_q:
        lecturers = query_db("SELECT * FROM users WHERE role = 'lecturer' AND (first_name LIKE ? OR last_name LIKE ? OR username LIKE ?) ORDER BY id DESC", ('%'+l_q+'%', '%'+l_q+'%', '%'+l_q+'%'))
    else:
        lecturers = query_db("SELECT * FROM users WHERE role = 'lecturer' ORDER BY id DESC")

    # Pass minimum data needed for the dashboard; others moved to Infrastructure page
    faculties = query_db('SELECT * FROM faculties')
    departments = query_db('SELECT * FROM departments')
    db_reset_enabled = True  # Always enabled as requested
    return render_template('admin.html', users=users, lecturers=lecturers, faculties=faculties, departments=departments, db_reset_enabled=db_reset_enabled, page=page, total_pages=total_pages, website_status=website_status, security_mode=security_mode)

@app.route('/admin/faculty/create', methods=['POST'])
def create_faculty():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    name = request.form.get('faculty_name')
    code = request.form.get('faculty_code')
    if not name:
        flash('Faculty name required')
        return redirect(url_for('superadmin_dashboard'))

    commit_db('INSERT INTO faculties (name, code) VALUES (?, ?) ON CONFLICT (name) DO NOTHING', (name, code))
    f_row = query_db('SELECT id FROM faculties WHERE name = ?', (name,), one=True)
    if f_row:
        log_action(g.user['id'], 'add', 'faculties', f_row['id'], None, {'name': name, 'code': code})
    flash('Faculty created')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/admin/department/create', methods=['POST'])
def create_department():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    name = request.form.get('department_name')
    code = request.form.get('department_code')
    faculty_id = request.form.get('faculty_id')

    if not name or not faculty_id:
        flash('Department name and Faculty required')
        return redirect(url_for('superadmin_dashboard'))

    commit_db('INSERT INTO departments (faculty_id, name, code) VALUES (?, ?, ?) ON CONFLICT DO NOTHING', (faculty_id, name, code))
    d_row = query_db('SELECT id FROM departments WHERE name = ? AND faculty_id = ?', (name, faculty_id), one=True)
    if d_row:
        log_action(g.user['id'], 'add', 'departments', d_row['id'], None, {'name': name, 'code': code, 'faculty_id': faculty_id})
    flash('Department created')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/student-profile/<int:student_id>')
def student_profile(student_id):
    if not g.user or g.user['role'] not in ['admin', 'lecturer', 'superadmin']:
        flash('Unauthorized access')
        return redirect(url_for('portal_hub'))

    student = query_db("SELECT * FROM users WHERE id = ? AND role = 'student'", (student_id,), one=True)
    if not student:
        flash('Student not found')
        return redirect(request.referrer or url_for('portal_hub'))

    # Calculate overall attendance
    total_sessions = query_db("SELECT COUNT(*) as c FROM sessions", one=True)['c']
    attended_sessions = query_db("SELECT COUNT(*) as c FROM attendance WHERE student_id = ?", (student_id,), one=True)['c']
    attendance_percent = round((attended_sessions / total_sessions * 100) if total_sessions > 0 else 0, 1)

    # Fetch department/faculty
    dept_info = None
    if student.get('department_id'):
        dept_info = query_db('''
            SELECT d.name as department_name, f.name as faculty_name
            FROM departments d
            JOIN faculties f ON d.faculty_id = f.id
            WHERE d.id = ?
        ''', (student['department_id'],), one=True)

    # Fetch recent attendance history
    history = query_db('''
        SELECT a.timestamp, a.status, s.course_name, s.start_time
        FROM attendance a
        JOIN sessions s ON a.session_id = s.id
        WHERE a.student_id = ?
        ORDER BY a.timestamp DESC
        LIMIT 10
    ''', (student_id,))

    # Fetch ID Card
    card = query_db("SELECT * FROM student_cards WHERE student_id = ?", (student_id,), one=True)

    return render_template('student_profile.html',
                           student=student,
                           total_sessions=total_sessions,
                           attended_sessions=attended_sessions,
                           attendance_percent=attendance_percent,
                           dept=dept_info,
                           history=history,
                           card=card)

@app.route('/admin/course/create', methods=['POST'])
def create_course():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    code = request.form.get('course_code', '').strip()
    title = request.form.get('course_title', '').strip()
    name = code or title
    if not name:
        flash('Course code or title required')
        return redirect(url_for('superadmin_dashboard'))
    commit_db('INSERT INTO courses (name, code, title) VALUES (?, ?, ?) ON CONFLICT (name) DO NOTHING', (name, code or None, title or None))
    c_row = query_db('SELECT id FROM courses WHERE name = ?', (name,), one=True)
    if c_row:
        log_action(g.user['id'], 'add', 'courses', c_row['id'], None, {'code': code, 'title': title})
    flash('Course created')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/admin/settings/update', methods=['POST'])
def update_settings():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))

    key = request.form.get('key')

    # Handle color and access updates (Superadmin only)
    if key in ('primary_color', 'secondary_color', 'github_access_for_admin'):
        if g.user['role'] != 'superadmin':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        value = request.form.get('value')
        if value:
            commit_db('INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value', (key, value))
            global _SETTINGS_CACHE_TIME
            _SETTINGS_CACHE_TIME = 0 # Invalidate cache
            return jsonify({'success': True, 'message': f"{key.replace('_', ' ').capitalize()} updated"})

    # Use 'logo' field name for background (legacy) and 'image' for others
    file_field = 'logo' if key == 'background' else 'image'

    if key in ('background', 'site_logo', 'favicon'):
        if file_field in request.files:
            file = request.files[file_field]
            if file and file.filename:
                img_data = file.read()
                if img_data:
                    b64_str = base64.b64encode(img_data).decode('utf-8')
                    mimetype = file.mimetype or 'image/png'
                    image_b64 = f"data:{mimetype};base64,{b64_str}"
                    commit_db('INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value', (key, image_b64))
                    flash(f"{key.replace('_', ' ').capitalize()} updated successfully")

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/upload-logo', methods=['POST'])
def upload_logo():
    # Keep legacy route but redirect to new one or handle as background
    return redirect(url_for('update_settings'))


@app.route('/lecturer/update')
def lecturer_update():
    try:
        return _lecturer_update_logic()
    except Exception as e:
        app.logger.error(f"Lecturer update error: {e}", exc_info=True)
        return f"Internal Server Error: {e}", 500

def _lecturer_update_logic():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    # Filters
    course_id = request.args.get('course_id')
    faculty_id = request.args.get('faculty_id')
    department_id = request.args.get('department_id')
    q = (request.args.get('q') or '').strip()
    start = request.args.get('start')
    end = request.args.get('end')

    # Determine students and attendance counts
    # Aggregate attendance per student across sessions for this lecturer

    # Pre-calculate total sessions for the lecturer to avoid nested subquery overhead
    total_params = [g.user['id']]
    total_sql = 'SELECT COUNT(*) as c FROM sessions WHERE lecturer_id = ?'
    if course_id:
        total_sql += ' AND course_id = ?'; total_params.append(course_id)
    if start:
        total_sql += ' AND start_time >= ?'; total_params.append(start)
    if end:
        total_sql += ' AND start_time <= ?'; total_params.append(end)

    total_res = query_db(total_sql, tuple(total_params), one=True)
    total_sessions_count = total_res['c'] if total_res else 0

    params = []
    sql = '''SELECT u.id as student_id,
               COALESCE(u.ub_id, '') as ub,
               COALESCE(u.matric, '') as matric,
               u.first_name, u.middle_name, u.last_name,
               COUNT(a.id) as present
             FROM users u
             LEFT JOIN attendance a ON a.student_id = u.id
             LEFT JOIN sessions ss ON a.session_id = ss.id
             WHERE u.role = 'student' AND (ss.lecturer_id = ? OR ss.lecturer_id IS NULL) '''
    params.append(g.user['id'])


    if course_id:
        sql += ' AND (ss.course_id = ? OR ss.course_id IS NULL)'
        params.append(course_id)
    if start:
        sql += ' AND (ss.start_time >= ? OR ss.start_time IS NULL)'
        params.append(start)
    if end:
        sql += ' AND (ss.start_time <= ? OR ss.start_time IS NULL)'
        params.append(end)
    if faculty_id:
        sql += ' AND (u.faculty_id = ? OR u.department_id IN (SELECT id FROM departments WHERE faculty_id = ?))'
        params.extend([faculty_id, faculty_id])
    if department_id:
        sql += ' AND u.department_id = ?'
        params.append(department_id)
    if q:
        sql += ' AND (u.first_name LIKE ? OR u.last_name LIKE ? OR u.ub_id LIKE ? OR u.matric LIKE ?)'
        params.extend(['%'+q+'%', '%'+q+'%', '%'+q+'%', '%'+q+'%'])

    final_sql = sql + ' GROUP BY u.id ORDER BY present DESC LIMIT 200'
    rows = query_db(final_sql, tuple(params))

    # Compute percent for each
    out = []
    for r in rows:
        total = total_sessions_count
        present = r['present'] or 0
        percent = round((present/total*100) if total>0 else 0,1)
        out.append({'student_id': r['student_id'], 'ub': r['ub'], 'matric': r['matric'], 'first_name': r['first_name'], 'middle_name': r['middle_name'], 'last_name': r['last_name'], 'present': present, 'total': total, 'percent': percent})

    if request.args.get('format') == 'json':
        return jsonify({'rows': out})

    courses = query_db('SELECT * FROM courses')
    faculties = query_db('SELECT * FROM faculties')
    departments = query_db('SELECT * FROM departments')
    return render_template('lecturer_update.html', rows=out, courses=courses, faculties=faculties, departments=departments)

@app.route('/statistics')
def statistics():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    # Filters for statistics
    filter_type = request.args.get('filter', 'daily')  # daily, monthly, yearly
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    sort_by = request.args.get('sort', 'attendance')  # attendance or name
    search = request.args.get('search', '').strip()

    # Get attendance data based on filter
    params = [g.user['id']]
    date_clause = ''
    if start_date:
        date_clause += ' AND a.timestamp >= ?'
        params.append(start_date)
    if end_date:
        date_clause += ' AND a.timestamp <= ?'
        params.append(end_date)

    if filter_type == 'daily':
        sql = '''SELECT DATE(a.timestamp) as date, COUNT(*) as count
                 FROM attendance a
                 JOIN sessions s ON a.session_id = s.id
                 WHERE s.lecturer_id = ?''' + date_clause + '''
                 GROUP BY DATE(a.timestamp)
                 ORDER BY date DESC'''
    elif filter_type == 'monthly':
        sql = '''SELECT to_char(a.timestamp, 'YYYY-MM') as month, COUNT(*) as count
                 FROM attendance a
                 JOIN sessions s ON a.session_id = s.id
                 WHERE s.lecturer_id = ?''' + date_clause + '''
                 GROUP BY to_char(a.timestamp, 'YYYY-MM')
                 ORDER BY month DESC'''
    else:  # yearly
        sql = '''SELECT to_char(a.timestamp, 'YYYY') as year, COUNT(*) as count
                 FROM attendance a
                 JOIN sessions s ON a.session_id = s.id
                 WHERE s.lecturer_id = ?''' + date_clause + '''
                 GROUP BY to_char(a.timestamp, 'YYYY')
                 ORDER BY year DESC'''

    stats = query_db(sql, tuple(params))

    # For student sorting and search
    if sort_by == 'attendance':
        order = 'present DESC'
    else:
        order = 'u.first_name ASC'

    search_clause = ''
    if search:
        search_clause = ''' AND (u.first_name LIKE ? OR u.last_name LIKE ? OR u.ub_id LIKE ?)'''
        params.extend(['%' + search + '%', '%' + search + '%', '%' + search + '%'])

    student_sql = '''SELECT u.id, u.first_name, u.last_name, u.ub_id, u.matric, COUNT(a.id) as present
                     FROM users u
                     LEFT JOIN attendance a ON u.id = a.student_id
                     LEFT JOIN sessions s ON a.session_id = s.id
                     WHERE u.role = 'student' AND s.lecturer_id = ?''' + date_clause + search_clause + '''
                     GROUP BY u.id
                     ORDER BY ''' + order
    students = query_db(student_sql, tuple(params))

    return render_template('statistics.html', stats=stats, students=students, filter_type=filter_type, start_date=start_date, end_date=end_date, sort_by=sort_by, search=search)

@app.route('/statistics/export')
def export_statistics():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    filter_type = request.args.get('filter', 'daily')
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    sort_by = request.args.get('sort', 'attendance')
    search = request.args.get('search', '').strip()

    # Similar logic as above to get data
    params = [g.user['id']]
    date_clause = ''
    if start_date:
        date_clause += ' AND a.timestamp >= ?'
        params.append(start_date)
    if end_date:
        date_clause += ' AND a.timestamp <= ?'
        params.append(end_date)

    if sort_by == 'attendance':
        order = 'present DESC'
    else:
        order = 'u.first_name ASC'

    search_clause = ''
    if search:
        search_clause = ''' AND (u.first_name LIKE ? OR u.last_name LIKE ? OR u.ub_id LIKE ?)'''
        params.extend(['%' + search + '%', '%' + search + '%', '%' + search + '%'])

    student_sql = '''SELECT u.first_name, u.last_name, u.ub_id, u.matric,
                            d.name as department_name, f.name as faculty_name,
                            COUNT(a.id) as present
                     FROM users u
                     LEFT JOIN departments d ON u.department_id = d.id
                     LEFT JOIN faculties f ON d.faculty_id = f.id
                     LEFT JOIN attendance a ON u.id = a.student_id
                     LEFT JOIN sessions s ON a.session_id = s.id
                     WHERE u.role = 'student' AND s.lecturer_id = ?''' + date_clause + search_clause + '''
                     GROUP BY u.id, d.name, f.name
                     ORDER BY ''' + order
    students = query_db(student_sql, tuple(params))

    # Export to Excel
    data = [{
        'First Name': s['first_name'],
        'Last Name': s['last_name'],
        'UB ID': s['ub_id'],
        'Matric': s['matric'],
        'Faculty': s['faculty_name'] or 'N/A',
        'Department': s['department_name'] or 'N/A',
        'Attendance Count': s['present']
    } for s in students]
    if pd:
        df = pd.DataFrame(data)
        bio = BytesIO()
        df.to_excel(bio, index=False, engine='openpyxl')
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name='attendance_statistics.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        # Fallback to CSV
        csv = 'First Name,Last Name,UB ID,Matric,Faculty,Department,Attendance Count\n'
        for row in data:
            csv += f"{row['First Name']},{row['Last Name']},{row['UB ID']},{row['Matric']},{row['Faculty']},{row['Department']},{row['Attendance Count']}\n"
        return send_file(BytesIO(csv.encode()), as_attachment=True, download_name='attendance_statistics.csv', mimetype='text/csv')

@app.route('/api/users/search')
def api_users_search():
    if not g.user:
        return jsonify({'error': 'Unauthorized'}), 403
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    # Search users by name, UB ID, or username
    users = query_db("SELECT id, first_name, last_name, ub_id, username, role FROM users WHERE (first_name LIKE ? OR last_name LIKE ? OR ub_id LIKE ? OR username LIKE ? OR matric LIKE ?) AND active = 1 AND id != ? LIMIT 15",
                        ('%' + q + '%', '%' + q + '%', '%' + q + '%', '%' + q + '%', '%' + q + '%', g.user['id']))
    results = []
    for u in users:
        identifier = u['ub_id'] if u['role'] == 'student' else u['username']
        results.append({
            'id': u['id'],
            'name': f"{u['first_name']} {u['last_name']}",
            'identifier': identifier,
            'role': u['role']
        })
    return jsonify(results)

@app.route('/api/students/search')
def api_students_search():
    if not g.user or g.user['role'] not in ['lecturer', 'admin', 'superadmin']:
        return jsonify({'error': 'Unauthorized'}), 403
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    # Search students by name or UB ID
    students = query_db("SELECT id, first_name, last_name, ub_id FROM users WHERE role = 'student' AND (first_name LIKE ? OR last_name LIKE ? OR ub_id LIKE ? OR matric LIKE ?) LIMIT 10",
                        ('%' + q + '%', '%' + q + '%', '%' + q + '%', '%' + q + '%'))
    results = [{'id': s['id'], 'name': f"{s['first_name']} {s['last_name']}", 'ub_id': s['ub_id']} for s in students]
    return jsonify(results)


@app.route('/superadmin')
def superadmin_dashboard():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    stats = {
        'students': query_db("SELECT COUNT(*) as c FROM users WHERE role = 'student'", one=True)['c'],
        'lecturers': query_db("SELECT COUNT(*) as c FROM users WHERE role = 'lecturer'", one=True)['c'],
        'courses': query_db("SELECT COUNT(*) as c FROM courses", one=True)['c'],
        'active_sessions': query_db("SELECT COUNT(*) as c FROM sessions WHERE active = 1", one=True)['c']
    }

    return render_template('superadmin_command_center.html', stats=stats)

@app.route('/superadmin/users')
def superadmin_users():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    per_page = 50
    offset = (page - 1) * per_page

    q = request.args.get('q', '').strip()
    faculty_id = request.args.get('faculty_id')
    department_id = request.args.get('department_id')

    query_base = "FROM users WHERE role = 'student'"
    params = []

    if q:
        query_base += " AND (first_name LIKE ? OR last_name LIKE ? OR ub_id LIKE ? OR matric LIKE ?)"
        params.extend(['%'+q+'%', '%'+q+'%', '%'+q+'%', '%'+q+'%'])

    if faculty_id:
        query_base += " AND (faculty_id = ? OR department_id IN (SELECT id FROM departments WHERE faculty_id = ?))"
        params.extend([faculty_id, faculty_id])

    if department_id:
        query_base += " AND department_id = ?"
        params.append(department_id)

    users = query_db(f"SELECT * {query_base} ORDER BY id DESC LIMIT ? OFFSET ?", tuple(params + [per_page, offset]))
    total_res = query_db(f"SELECT COUNT(*) as c {query_base}", tuple(params), one=True)
    total_users = total_res['c'] if total_res else 0
    total_pages = (total_users + per_page - 1) // per_page

    l_q = request.args.get('l_q', '').strip()
    if l_q:
        lecturers = query_db("SELECT * FROM users WHERE role = 'lecturer' AND (first_name LIKE ? OR last_name LIKE ? OR username LIKE ?) ORDER BY id DESC", ('%'+l_q+'%', '%'+l_q+'%', '%'+l_q+'%'))
    else:
        lecturers = query_db("SELECT * FROM users WHERE role = 'lecturer' ORDER BY id DESC")

    faculties = query_db('SELECT * FROM faculties')
    departments = query_db('SELECT * FROM departments')

    return render_template('superadmin_users.html',
                           users=users, lecturers=lecturers, faculties=faculties, departments=departments,
                           page=page, total_pages=total_pages)

@app.route('/superadmin/governance')
def superadmin_governance():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    website_status = get_setting('website_status', 'on')
    security_mode = get_setting('security_mode', 'off')
    strict_integrity = get_setting('strict_integrity_mode', 'off')

    return render_template('superadmin_governance.html',
                           website_status=website_status,
                           security_mode=security_mode,
                           strict_integrity=strict_integrity)

@app.route('/superadmin/settings')
def superadmin_settings():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    return render_template('superadmin_settings.html')

@app.route('/superadmin/team')
def superadmin_team():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    members = get_team_members_with_fallback(is_admin=True)
    roles = query_db('SELECT * FROM team_roles')
    return render_template('superadmin_team.html', members=members, roles=roles)

@app.route('/superadmin/finance')
def superadmin_finance():
    if not g.user or g.user['role'] != 'superadmin':
        return redirect(url_for('login', role='superadmin'))

    student_cost = float(get_setting('student_cost', 0))
    billing_period = get_setting('billing_period', 'session')

    student_count = query_db("SELECT COUNT(*) as c FROM users WHERE role = 'student'", one=True)['c']
    total_revenue = student_count * student_cost

    return render_template('superadmin_finance.html',
                           student_count=student_count,
                           student_cost=student_cost,
                           billing_period=billing_period,
                           total_revenue=total_revenue)

@app.route('/superadmin/auto-seed', methods=['GET', 'POST'])
@csrf.exempt
def superadmin_auto_seed():
    if not g.user or g.user['role'] != 'superadmin':
        if request.method == 'POST':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        return redirect(url_for('login', role='superadmin'))

    if request.method == 'GET':
        return render_template('superadmin_auto_seed.html')

    results = {'faculties': 0, 'departments': 0, 'courses': 0, 'source': 'unknown'}

    def load_from_static():
        path = os.path.join(os.path.dirname(__file__), 'seed_data_uniben.json')
        if not os.path.exists(path):
            raise FileNotFoundError("seed_data_uniben.json not found")
        with open(path, encoding='utf-8') as f:
            return json.load(f)

    def seed_faculty(fac):
        f = query_db("SELECT id, name FROM faculties WHERE code = ?", (fac['code'],), one=True)
        if f:
            if f['name'] != fac['name']:
                commit_db("UPDATE faculties SET name = ? WHERE id = ?", (fac['name'], f['id']))
            return f['id']
        f_by_name = query_db("SELECT id, code FROM faculties WHERE name = ?", (fac['name'],), one=True)
        if f_by_name:
            if f_by_name['code'] != fac['code']:
                commit_db("UPDATE faculties SET code = ? WHERE id = ?", (fac['code'], f_by_name['id']))
            return f_by_name['id']
        commit_db("INSERT INTO faculties (name, code) VALUES (?, ?)", (fac['name'], fac['code']))
        f_new = query_db("SELECT id FROM faculties WHERE code = ?", (fac['code'],), one=True)
        return f_new['id'] if f_new else None

    def seed_dept(f_id, dept):
        d = query_db("SELECT id, name FROM departments WHERE faculty_id = ? AND code = ?", (f_id, dept['code']), one=True)
        if d:
            if d['name'] != dept['name']:
                commit_db("UPDATE departments SET name = ? WHERE id = ?", (dept['name'], d['id']))
            return d['id']
        d_by_name = query_db("SELECT id, code FROM departments WHERE faculty_id = ? AND name = ?", (f_id, dept['name']), one=True)
        if d_by_name:
            if d_by_name['code'] != dept['code']:
                commit_db("UPDATE departments SET code = ? WHERE id = ?", (dept['code'], d_by_name['id']))
            return d_by_name['id']
        commit_db("INSERT INTO departments (faculty_id, name, code) VALUES (?, ?, ?)", (f_id, dept['name'], dept['code']))
        d_new = query_db("SELECT id FROM departments WHERE faculty_id = ? AND code = ?", (f_id, dept['code']), one=True)
        return d_new['id'] if d_new else None

    def seed_course(code, title):
        if not code:
            return
        commit_db(
            'INSERT INTO courses (name, code, title) VALUES (?, ?, ?) ON CONFLICT (name) DO UPDATE SET code = EXCLUDED.code, title = EXCLUDED.title',
            (code, code, title or '')
        )

    try:
        try:
            base_urls = ["https://waeup.uniben.edu", "https://uniben.waeup.org"]
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
            }

            def try_fetch(url, timeout=15):
                for bu in base_urls:
                    try:
                        resp = requests.get(f"{bu}{url}", timeout=timeout, headers=headers)
                        resp.raise_for_status()
                        return resp
                    except requests.RequestException:
                        continue
                return None

            resp = try_fetch("/faculties")
            if resp:
                results['source'] = 'live'
                soup = BeautifulSoup(resp.text, 'html.parser')
                faculties = []
                for row in soup.find_all('tr'):
                    cols = row.find_all('td')
                    if len(cols) >= 2:
                        a = cols[0].find('a')
                        if a and '/faculties/' in a['href']:
                            faculties.append({'code': a['href'].split('/')[-1], 'name': cols[1].get_text(strip=True)})

                results['faculties'] = len(faculties)

                for fac in faculties:
                    f_id = seed_faculty(fac)
                    if not f_id:
                        continue
                    dept_resp = try_fetch(f"/faculties/{fac['code']}")
                    if not dept_resp:
                        continue
                    dept_soup = BeautifulSoup(dept_resp.text, 'html.parser')
                    departments = []
                    for row in dept_soup.find_all('tr'):
                        cols = row.find_all('td')
                        if len(cols) >= 2:
                            a = cols[0].find('a')
                            if a and f'/faculties/{fac["code"]}/' in a['href']:
                                departments.append({'code': a['href'].split('/')[-1], 'name': cols[1].get_text(strip=True)})

                    results['departments'] += len(departments)

                    for dept in departments:
                        d_id = seed_dept(f_id, dept)
                        if not d_id:
                            continue
                        course_resp = try_fetch(f"/faculties/{fac['code']}/{dept['code']}")
                        if not course_resp:
                            continue
                        course_soup = BeautifulSoup(course_resp.text, 'html.parser')
                        for row in course_soup.find_all('tr'):
                            cols = row.find_all('td')
                            if len(cols) >= 2:
                                a = cols[0].find('a')
                                if a and '/courses/' in a['href']:
                                    seed_course(a.get_text(strip=True), cols[1].get_text(strip=True))
                                    results['courses'] += 1

        except Exception:
            data = load_from_static()
            results['source'] = 'static'
            for fac in data['faculties']:
                f_id = seed_faculty(fac)
                if not f_id:
                    continue
                results['faculties'] += 1
                for dept in fac.get('departments', []):
                    d_id = seed_dept(f_id, dept)
                    if not d_id:
                        continue
                    results['departments'] += 1
                    for course in dept.get('courses', []):
                        seed_course(course.get('code', ''), course.get('title', ''))
                        results['courses'] += 1

        return jsonify({
            'success': True,
            'message': f"Auto-seed complete ({results['source']}): {results['faculties']} faculties, {results['departments']} departments, {results['courses']} courses",
            **results
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/superadmin/finance/update', methods=['POST'])
def update_finance_settings():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    cost = request.form.get('student_cost')
    period = request.form.get('billing_period')

    if cost is not None:
        commit_db("INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", ('student_cost', cost))
    if period is not None:
        commit_db("INSERT INTO settings (key, value) VALUES (?,?) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", ('billing_period', period))

    global _SETTINGS_CACHE_TIME
    _SETTINGS_CACHE_TIME = 0

    flash('Financial settings updated successfully')
    return redirect(url_for('superadmin_finance'))


@app.route('/lecturer/student-attendance/<int:student_id>')
def lecturer_student_attendance(student_id):
    # Return JSON timeseries for student attendance for charting
    course_id = request.args.get('course_id')
    start = request.args.get('start')
    end = request.args.get('end')
    # Find sessions for this lecturer
    params = [g.user['id']]
    sql = 'SELECT id, start_time FROM sessions WHERE lecturer_id = ?'
    if course_id:
        sql += ' AND course_id = ?'; params.append(course_id)
    if start:
        sql += ' AND start_time >= ?'; params.append(start)
    if end:
        sql += ' AND start_time <= ?'; params.append(end)
    sessions = query_db(sql, tuple(params))
    labels = []
    counts = []
    for s in sessions:
        st = s['start_time']
        if isinstance(st, datetime):
            st = st.strftime('%Y-%m-%d %H:%M:%S')
        labels.append(st)
        c = query_db('SELECT COUNT(*) as c FROM attendance WHERE session_id = ? AND student_id = ?', (s['id'], student_id), one=True)['c']
        counts.append(c)
    return jsonify({'labels': labels, 'counts': counts})




@app.route('/admin/faculty/delete/<int:faculty_id>', methods=['POST'])
def delete_faculty(faculty_id):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    old_f = query_db("SELECT * FROM faculties WHERE id = ?", (faculty_id,), one=True)
    if not old_f:
        return jsonify({'success': False, 'message': 'Faculty not found'}), 404

    try:
        # Log action before deletion
        log_action(g.user['id'], 'delete', 'faculties', faculty_id, old_f, None)

        # Cascaded Cleanup: Users and Departments
        commit_db("UPDATE users SET faculty_id = NULL WHERE faculty_id = ?", (faculty_id,))
        # Find all departments belonging to this faculty and set their users' faculty/dept/assigned_dept to NULL
        depts = query_db("SELECT id FROM departments WHERE faculty_id = ?", (faculty_id,))
        for d in depts:
            commit_db("UPDATE users SET department_id = NULL, assigned_dept_id = NULL WHERE department_id = ? OR assigned_dept_id = ?", (d['id'], d['id']))

        commit_db("DELETE FROM departments WHERE faculty_id = ?", (faculty_id,))
        commit_db("DELETE FROM faculties WHERE id = ?", (faculty_id,))

        security_log("FACULTY_DELETED", f"Faculty {faculty_id} ({old_f['name']}) deleted by {g.user['role']} {g.user['id']}")
        return jsonify({'success': True, 'message': 'Faculty and linked departments removed.'})
    except Exception as e:
        app.logger.error(f"Error deleting faculty {faculty_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/department/delete/<int:department_id>', methods=['POST'])
def delete_department(department_id):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    old_d = query_db("SELECT * FROM departments WHERE id = ?", (department_id,), one=True)
    if not old_d:
        return jsonify({'success': False, 'message': 'Department not found'}), 404

    try:
        # Log action
        log_action(g.user['id'], 'delete', 'departments', department_id, old_d, None)

        # Cleanup: Users
        commit_db("UPDATE users SET department_id = NULL, assigned_dept_id = NULL WHERE department_id = ? OR assigned_dept_id = ?", (department_id, department_id))

        commit_db("DELETE FROM departments WHERE id = ?", (department_id,))

        security_log("DEPARTMENT_DELETED", f"Department {department_id} ({old_d['name']}) deleted by {g.user['role']} {g.user['id']}")
        return jsonify({'success': True, 'message': 'Department removed successfully.'})
    except Exception as e:
        app.logger.error(f"Error deleting department {department_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/course/delete/<int:course_id>', methods=['POST'])
def delete_course_route(course_id):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    old_c = query_db("SELECT * FROM courses WHERE id = ?", (course_id,), one=True)
    if not old_c:
        return jsonify({'success': False, 'message': 'Course not found'}), 404

    try:
        # Log action
        log_action(g.user['id'], 'delete', 'courses', course_id, old_c, None)

        # Cleanup: Lecturer courses and sessions
        commit_db("DELETE FROM lecturer_courses WHERE course_id = ?", (course_id,))
        commit_db("UPDATE sessions SET course_id = NULL WHERE course_id = ?", (course_id,))

        commit_db("DELETE FROM courses WHERE id = ?", (course_id,))

        security_log("COURSE_DELETED", f"Course {course_id} ({old_c['name']}) deleted by {g.user['role']} {g.user['id']}")
        return jsonify({'success': True, 'message': 'Course removed from catalog.'})
    except Exception as e:
        app.logger.error(f"Error deleting course {course_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/departments/<int:faculty_id>')
def api_departments(faculty_id):
    deps = query_db('SELECT id, name FROM departments WHERE faculty_id = ?', (faculty_id,))
    out = [{'id': d['id'], 'name': d['name']} for d in deps]
    return jsonify(out)

@app.route('/api/faculties')
def api_all_faculties():
    rows = query_db('SELECT id, name FROM faculties ORDER BY name')
    return jsonify([dict(r) for r in rows])

@app.route('/api/departments')
def api_all_departments():
    rows = query_db('SELECT id, name FROM departments ORDER BY name')
    return jsonify([dict(r) for r in rows])

@app.route('/api/announcements/send', methods=['POST'])
@limiter.limit("30 per minute")
def send_announcement():
    if not g.user or g.user['role'] not in ['lecturer', 'admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    target_type = data.get('target_type')
    target_id = data.get('target_id')
    message = data.get('message', '').strip()
    priority = data.get('priority', 'normal')

    if not message:
        return jsonify({'success': False, 'message': 'Message cannot be empty'}), 400
    if target_type not in ['student', 'students', 'lecturers', 'department', 'faculty', 'system']:
        return jsonify({'success': False, 'message': 'Invalid target type'}), 400

    # Ensure target_id is integer if provided
    if target_id is not None:
        try:
            target_id = int(target_id)
        except (ValueError, TypeError):
            target_id = None

    commit_db("INSERT INTO announcements (sender_id, target_type, target_id, message, priority) VALUES (?,?,?,?,?)",
              (g.user['id'], target_type, target_id, message, priority))

    return jsonify({'success': True, 'message': 'Announcement broadcasted successfully'})

@app.route('/announcements')
def announcements_page():
    if not g.user:
        return redirect(url_for('login', role='student'))
    return render_template('announcements.html')


@app.route('/api/announcements/my')
def my_announcements():
    if not g.user:
        return jsonify([])

    # Determine department and faculty for the current user to filter announcements
    dept_id = g.user.get('department_id') or g.user.get('assigned_dept_id')
    faculty_id = g.user.get('faculty_id')

    if dept_id and not faculty_id:
        dept = query_db("SELECT faculty_id FROM departments WHERE id = ?", (dept_id,), one=True)
        if dept:
            faculty_id = dept['faculty_id']

    query = """
        SELECT a.*, u.first_name, u.last_name, u.role as sender_role
        FROM announcements a
        LEFT JOIN users u ON a.sender_id = u.id
        WHERE a.target_type = 'system'
    """
    params = []

    if g.user['role'] == 'student':
        query += " OR (a.target_type = 'student' AND a.target_id = ?)"
        params.append(g.user['id'])
        query += " OR a.target_type = 'students'"
    elif g.user['role'] == 'lecturer':
        query += " OR a.target_type = 'lecturers'"

    if dept_id:
        query += " OR (a.target_type = 'department' AND a.target_id = ?)"
        params.append(dept_id)

    if faculty_id:
        query += " OR (a.target_type = 'faculty' AND a.target_id = ?)"
        params.append(faculty_id)

    query += " ORDER BY a.created_at DESC LIMIT 20"

    rows = query_db(query, tuple(params))

    # Convert created_at to string for JSON
    out = []
    for r in rows:
        r_dict = dict(r)
        if isinstance(r_dict['created_at'], datetime):
            r_dict['created_at'] = r_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        # Handle potentially missing user data from LEFT JOIN
        r_dict['first_name'] = r_dict.get('first_name') or 'System'
        r_dict['last_name'] = r_dict.get('last_name') or 'Admin'
        r_dict['sender_role'] = r_dict.get('sender_role') or 'admin'
        out.append(r_dict)

    return jsonify(out)


@app.route('/admin/toggle-website', methods=['POST'])
def toggle_website():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    global _SETTINGS_CACHE_TIME
    current = get_setting('website_status', 'on')
    new_status = 'off' if current == 'on' else 'on'
    commit_db("UPDATE settings SET value = ? WHERE key = 'website_status'", (new_status,))
    _SETTINGS_CACHE_TIME = 0 # Invalidate cache
    return jsonify({'success': True, 'new_status': new_status})


@app.route('/admin/toggle-security', methods=['POST'])
def toggle_security():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    global _SETTINGS_CACHE_TIME
    current = get_setting('security_mode', 'off')
    new_status = 'off' if current == 'on' else 'on'
    commit_db("UPDATE settings SET value = ? WHERE key = 'security_mode'", (new_status,))
    _SETTINGS_CACHE_TIME = 0 # Invalidate cache

    # If turning off, clear blocks immediately
    if new_status == 'off':
        commit_db("DELETE FROM blocked_ips")
        commit_db("DELETE FROM failed_logins")

    return jsonify({'success': True, 'new_status': new_status})

@app.route('/admin/toggle-integrity', methods=['POST'])
def toggle_integrity():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    global _SETTINGS_CACHE_TIME
    current = get_setting('strict_integrity_mode', 'off')
    new_status = 'off' if current == 'on' else 'on'
    commit_db("UPDATE settings SET value = ? WHERE key = 'strict_integrity_mode'", (new_status,))
    _SETTINGS_CACHE_TIME = 0 # Invalidate cache
    return jsonify({'success': True, 'new_status': new_status})

@app.route('/admin/shutdown-role', methods=['POST'])
def shutdown_role():
    if not g.user or g.user['role'] != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    role = data.get('role')
    action = data.get('action') # 'shutdown' or 'activate'

    if role not in ['student', 'lecturer', 'admin']:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400

    new_status = 0 if action == 'shutdown' else 1

    # Log the bulk action
    log_action(g.user['id'], 'bulk_toggle', 'users', -1, {'role': role, 'old_status': 1-new_status}, {'role': role, 'new_status': new_status})

    commit_db("UPDATE users SET active = ? WHERE role = ?", (new_status, role))

    security_log("ROLE_SHUTDOWN" if new_status == 0 else "ROLE_ACTIVATION", f"All users with role {role} set to active={new_status}")

    return jsonify({'success': True, 'message': f"All {role}s have been {'deactivated' if new_status == 0 else 'activated'}."})

@app.route('/admin/unblock-ip/<ip>', methods=['POST'])
def unblock_ip(ip):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    commit_db("DELETE FROM blocked_ips WHERE ip = ?", (ip,))
    commit_db("DELETE FROM failed_logins WHERE ip = ?", (ip,))
    return jsonify({'success': True, 'message': f'IP {ip} unblocked'})

@app.route('/admin/developer-settings', methods=['POST'])
def update_developer_settings():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    gh_token = data.get('github_token')
    v_token = data.get('vercel_token')
    v_project_id = data.get('vercel_project_id')

    if g.user['role'] == 'admin':
        status = get_setting('github_access_for_admin', 'off')
        if status == 'off' and gh_token and gh_token != g.user.get('github_token'):
             return jsonify({'success': False, 'message': 'GitHub linkage disabled for Admins.'}), 403

    commit_db("UPDATE users SET github_token = ?, vercel_token = ?, vercel_project_id = ? WHERE id = ?",
              (gh_token, v_token, v_project_id, g.user['id']))
    return jsonify({'success': True, 'message': 'Settings updated'})

@app.route('/admin/documentation')
def documentation_workspace():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    if g.user['role'] == 'admin' and get_setting('github_access_for_admin', 'off') == 'off':
        flash('GitHub Sync is disabled for Admins.'); return redirect(url_for('admin_dashboard'))
    return render_template('documentation.html')

@app.route('/admin/github/sync', methods=['POST'])
def github_sync_doc():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    if g.user['role'] == 'admin' and get_setting('github_access_for_admin', 'off') == 'off':
        return jsonify({'success': False, 'message': 'Access disabled'}), 403
    token = g.user.get('github_token')
    if not token: return jsonify({'success': False, 'message': 'GitHub token not set.'}), 400
    import requests
    data = request.get_json()
    repo = data.get('repo'); path = data.get('path', 'DOCS.md'); content = data.get('content')
    headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
    try:
        u_resp = requests.get("https://api.github.com/user", headers=headers)
        if u_resp.status_code != 200: return jsonify({'success': False, 'message': 'Invalid token'}), 401
        uname = u_resp.json()['login']
        f_repo = f"{uname}/{repo}"
        r_resp = requests.get(f"https://api.github.com/repos/{f_repo}", headers=headers)
        if r_resp.status_code == 404:
            requests.post("https://api.github.com/user/repos", headers=headers, json={"name": repo, "auto_init": True})
            time.sleep(2)
        f_url = f"https://api.github.com/repos/{f_repo}/contents/{path}"
        f_resp = requests.get(f_url, headers=headers)
        sha = f_resp.json()['sha'] if f_resp.status_code == 200 else None
        c_data = {"message": "Update docs via AtenLana", "content": base64.b64encode(content.encode()).decode(), "branch": "main"}
        if sha: c_data["sha"] = sha
        p_resp = requests.put(f_url, headers=headers, json=c_data)
        return jsonify({'success': True}) if p_resp.status_code in [200, 201] else jsonify({'success': False, 'message': 'Push failed'})
    except Exception as e: return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/sync/discover-faculties', methods=['POST'])
def sync_discover_faculties():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    base_url = "https://waeup.uniben.edu/faculties"
    try:
        resp = requests.get(base_url, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')
        faculties = []
        for row in soup.find_all('tr'):
            cols = row.find_all('td')
            if len(cols) >= 2:
                a = cols[0].find('a')
                if a and '/faculties/' in a['href']:
                    code = a['href'].split('/')[-1]
                    name = cols[1].get_text(strip=True)
                    faculties.append({'code': code, 'name': name})
        return jsonify({'success': True, 'faculties': faculties})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/sync/faculty-departments', methods=['POST'])
def sync_faculty_departments():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    code = data.get('code')
    name = data.get('name')
    if not code or not name: return jsonify({'success': False, 'message': 'Missing data'}), 400

    try:
        # 1. Ensure Faculty exists
        # Search by code first
        f = query_db("SELECT id, name FROM faculties WHERE code = ?", (code,), one=True)
        if f:
            # Overwrite name if it changed
            if f['name'] != name:
                commit_db("UPDATE faculties SET name = ? WHERE id = ?", (name, f['id']))
            f_id = f['id']
        else:
            # Fallback to name-based lookup for existing manual entries
            f_by_name = query_db("SELECT id, code FROM faculties WHERE name = ?", (name,), one=True)
            if f_by_name:
                # Update code if it was missing or different
                if f_by_name['code'] != code:
                    commit_db("UPDATE faculties SET code = ? WHERE id = ?", (code, f_by_name['id']))
                f_id = f_by_name['id']
            else:
                commit_db("INSERT INTO faculties (name, code) VALUES (?, ?)", (name, code))
                f_new = query_db("SELECT id FROM faculties WHERE code = ?", (code,), one=True)
                f_id = f_new['id'] if f_new else None

        if not f_id: return jsonify({'success': False, 'message': 'Failed to create faculty record'}), 500

        # 2. Scrape Departments
        dept_url = f"https://waeup.uniben.edu/faculties/{code}"
        resp = requests.get(dept_url, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')
        synced_depts = []
        for row in soup.find_all('tr'):
            cols = row.find_all('td')
            if len(cols) >= 2:
                a = cols[0].find('a')
                if a and f'/faculties/{code}/' in a['href']:
                    d_code = a['href'].split('/')[-1]
                    d_name = cols[1].get_text(strip=True)

                    # 1. Try code-based lookup
                    d = query_db("SELECT id, name FROM departments WHERE faculty_id = ? AND code = ?", (f_id, d_code), one=True)
                    if d:
                        # Overwrite name if it changed
                        if d['name'] != d_name:
                            commit_db("UPDATE departments SET name = ? WHERE id = ?", (d_name, d['id']))
                    else:
                        # 2. Try name-based lookup within faculty
                        d_by_name = query_db("SELECT id, code FROM departments WHERE faculty_id = ? AND name = ?", (f_id, d_name), one=True)
                        if d_by_name:
                            # Update code if it was missing or different
                            if d_by_name['code'] != d_code:
                                commit_db("UPDATE departments SET code = ? WHERE id = ?", (d_code, d_by_name['id']))
                        else:
                            # 3. New record
                            commit_db("INSERT INTO departments (faculty_id, name, code) VALUES (?, ?, ?)", (f_id, d_name, d_code))

                    synced_depts.append({'code': d_code, 'name': d_name})

        return jsonify({'success': True, 'count': len(synced_depts), 'departments': synced_depts})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/sync/department-courses', methods=['POST'])
def sync_department_courses():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    f_code = data.get('f_code')
    d_code = data.get('d_code')
    if not f_code or not d_code: return jsonify({'success': False, 'message': 'Missing data'}), 400

    try:
        url = f"https://waeup.uniben.edu/faculties/{f_code}/{d_code}"
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')

        synced = 0
        for a in soup.find_all('a', href=re.compile(r'/courses/')):
            course_name = a.get_text(strip=True)
            if course_name:
                commit_db('INSERT INTO courses (name) VALUES (?) ON CONFLICT (name) DO NOTHING', (course_name,))
                synced += 1

        return jsonify({'success': True, 'count': synced})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/bulk-generate-users', methods=['POST'])
def admin_bulk_generate_users():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    count = data.get('count', 1)
    role = data.get('role', 'student')
    faculty_id = data.get('faculty_id') or None
    dept_id = data.get('dept_id') or None

    try:
        count = int(count)
        if count <= 0 or count > 500:
            return jsonify({'success': False, 'message': 'Count must be between 1 and 500'}), 400
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Invalid count'}), 400

    first_names = ["John", "Jane", "Alice", "Bob", "Charlie", "David", "Eve", "Frank", "Grace", "Heidi", "Ivan", "Judy", "Kevin", "Laura", "Mallory", "Niaj", "Oscar", "Peggy", "Rupert", "Sybil", "Trent", "Victor", "Walter"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"]

    generated = 0
    for _ in range(count):
        fname = random.choice(first_names)
        lname = random.choice(last_names)
        pw_hash = generate_password_hash('password123')

        if role == 'student':
            ub_id = f"UB{random.randint(100000, 999999)}"
            matric = f"{random.choice(['CSC', 'MTH', 'PHY', 'CHM'])}/{random.randint(2020, 2024)}/{random.randint(100, 999)}"
            # Avoid duplicate UB IDs
            exists = query_db("SELECT id FROM users WHERE ub_id = ?", (ub_id,), one=True)
            if exists: continue

            commit_db("INSERT INTO users (role, first_name, last_name, ub_id, matric, password_hash, faculty_id, department_id, active) VALUES (?,?,?,?,?,?,?,?,?)",
                      ('student', fname, lname, ub_id, matric, pw_hash, faculty_id, dept_id, 1))
        elif role == 'lecturer':
            username = f"{fname.lower()}.{lname.lower()}{random.randint(10, 99)}"
            exists = query_db("SELECT id FROM users WHERE username = ?", (username,), one=True)
            if exists: continue

            commit_db("INSERT INTO users (role, username, first_name, last_name, password_hash, active) VALUES (?,?,?,?,?,?)",
                      ('lecturer', username, fname, lname, pw_hash, 1))
        elif role == 'admin':
            username = f"admin.{fname.lower()}{random.randint(10, 99)}"
            exists = query_db("SELECT id FROM users WHERE username = ?", (username,), one=True)
            if exists: continue
            commit_db("INSERT INTO users (role, username, first_name, last_name, password_hash, active) VALUES (?,?,?,?,?,?)",
                      ('admin', username, fname, lname, pw_hash, 1))

        generated += 1

    log_action(g.user['id'], 'bulk_add', 'users', -1, None, {'count': generated, 'role': role})
    return jsonify({'success': True, 'message': f'Successfully generated {generated} {role}s.'})

@app.route('/admin/sync-status')
def admin_sync_status():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    return jsonify(_SYNC_STATUS)

@app.route('/admin/update-role', methods=['POST'])
def update_user_role():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json()
    user_id = data.get('user_id')
    new_role = data.get('role')
    faculty_id = data.get('faculty_id')
    dept_id = data.get('dept_id')

    if not user_id or not new_role:
        return jsonify({'success': False, 'message': 'Missing data'}), 400

    old_user = query_db("SELECT * FROM users WHERE id = ?", (user_id,), one=True)

    # Validate role
    if new_role not in ['student', 'lecturer', 'admin']:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400

    # Update database
    commit_db("UPDATE users SET role = ?, faculty_id = ?, assigned_dept_id = ? WHERE id = ?",
              (new_role, faculty_id, dept_id, user_id))

    if old_user:
        old_data = {k: v for k, v in old_user.items() if k in ('role', 'faculty_id', 'assigned_dept_id')}
        new_data = {'role': new_role, 'faculty_id': faculty_id, 'assigned_dept_id': dept_id}
        log_action(g.user['id'], 'edit', 'users', user_id, old_data, new_data)

    security_log("ROLE_UPDATE", f"User {user_id} role updated to {new_role}")
    return jsonify({'success': True, 'message': 'User role and assignments updated'})

@app.route('/admin/delete/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    # Audit log
    try:
        app.logger.info(f"ADMIN_AUDIT: {datetime.now(timezone.utc).isoformat()} - admin {g.user['id']} deleted user {user_id}")
    except Exception:
        pass

    old_user = query_db("SELECT * FROM users WHERE id = ?", (user_id,), one=True)
    if old_user:
        attendance = query_db("SELECT * FROM attendance WHERE student_id = ?", (user_id,))
        # Convert objects to strings for JSON
        for att in attendance:
            for k, v in att.items():
                if isinstance(v, datetime):
                    att[k] = v.strftime('%Y-%m-%d %H:%M:%S')

        courses = []
        if old_user['role'] == 'lecturer':
            courses = query_db("SELECT * FROM lecturer_courses WHERE lecturer_id = ?", (user_id,))

        log_action(g.user['id'], 'delete', 'users', user_id, {'user': old_user, 'attendance': attendance, 'courses': courses}, None)

    commit_db('DELETE FROM users WHERE id = ?', (user_id,))
    # Also delete their attendance records/lecturer courses to be thorough
    commit_db('DELETE FROM attendance WHERE student_id = ?', (user_id,))
    commit_db('DELETE FROM lecturer_courses WHERE lecturer_id = ?', (user_id,))

    return jsonify({'success': True})

@app.route('/admin/toggle/<int:user_id>')
def toggle_user(user_id):
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    user = query_db('SELECT * FROM users WHERE id = ?', (user_id,), one=True)
    if not user: return 'User not found', 404
    new = 0 if user['active'] else 1
    commit_db('UPDATE users SET active = ? WHERE id = ?', (new, user_id))
    log_action(g.user['id'], 'toggle', 'users', user_id, {'active': user['active']}, {'active': new})

    # Redirect back to the referrer if possible, otherwise default to admin dashboard
    target = request.referrer
    if not target or 'toggle' in target: # Avoid redirect loops
        target = url_for('admin_dashboard')
    return redirect(target)

@app.route('/admin/reset', methods=['POST'])
def reset_database():
    """Danger action: Clear all tables and re-run init_db().
    Requires the exact confirmation phrase: "RESET DATABASE"."""
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))

    confirm = (request.form.get('confirm') or '').strip()
    if confirm != 'RESET DATABASE':
        flash('Confirmation phrase did not match. Type "RESET DATABASE" to confirm.')
        return redirect(url_for('admin_dashboard'))
    try:
        db = get_db()
        with db.cursor() as cur:
            # Drop all tables in the public schema (PostgreSQL)
            cur.execute("""
                DO $$ DECLARE
                    r RECORD;
                BEGIN
                    FOR r IN (SELECT tablename FROM pg_tables WHERE schemaname = 'public') LOOP
                        EXECUTE 'DROP TABLE IF EXISTS ' || quote_ident(r.tablename) || ' CASCADE';
                    END LOOP;
                END $$;
            """)
            db.commit()

        with app.app_context():
            init_db()

        # audit log
        try:
            app.logger.info(f"ADMIN_AUDIT: {datetime.now(timezone.utc).isoformat()} - admin {g.user['id']} ({g.user.get('first_name')}) reset the database")
        except Exception:
            pass
        flash('Database has been reset and re-initialized.')
    except Exception as e:
        flash('Database reset failed: ' + str(e))
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/purge-samples', methods=['POST'])
def purge_samples():
    """Remove example/sample users and courses if enabled via PURGE_SAMPLE=1."""
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    if os.environ.get('PURGE_SAMPLE','0') != '1':
        flash('Sample purge is disabled. Set PURGE_SAMPLE=1 to enable.')
        return redirect(url_for('admin_dashboard'))
    try:
        # Delete popularly seeded sample accounts and example courses
        sample_usernames = ('lecturer1','student1','student2','admin')
        for u in sample_usernames:
            commit_db('DELETE FROM users WHERE username = ?', (u,))
        sample_courses = ('CSC 101','MTH 201','PHY 102')
        for c in sample_courses:
            commit_db('DELETE FROM courses WHERE name = ?', (c,))
        flash('Sample data purged where present.')
    except Exception as e:
        flash('Purge failed: ' + str(e))
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/bulk-activate', methods=['POST'])
def admin_bulk_activate():
    if not g.user or g.user['role'] not in ['admin', 'superadmin']:
        return redirect(url_for('login', role='admin'))
    idents = (request.form.get('idents') or '').strip().splitlines()
    id_type = request.form.get('id_type')
    action = request.form.get('action')
    if not idents:
        flash('No identifiers provided')
        return redirect(url_for('admin_dashboard'))
    new_val = 1 if action == 'activate' else 0
    try:
        for it in idents:
            it = it.strip()
            if not it: continue
            if id_type == 'ub' or id_type == 'both':
                commit_db('UPDATE users SET active = ? WHERE ub_id = ?', (new_val, it))
            if id_type == 'matric' or id_type == 'both':
                commit_db('UPDATE users SET active = ? WHERE matric = ?', (new_val, it))
        flash('Bulk activation changes applied')
    except Exception as e:
        flash('Bulk operation failed: ' + str(e))
    return redirect(url_for('admin_dashboard'))

# Simple route to add saved location for lecturer
@app.route('/locations/add', methods=['POST'])
def add_location():
    if not g.user or g.user['role'] != 'lecturer':
        return redirect(url_for('login', role='lecturer'))
    name = request.form['name']
    lat = float(request.form['lat'])
    lon = float(request.form['lon'])
    commit_db('INSERT INTO locations (lecturer_id, name, latitude, longitude) VALUES (?,?,?,?)', (g.user['id'], name, lat, lon))
    return redirect(url_for('lecturer_dashboard'))

@app.route('/locations/update-current', methods=['POST'])
def update_current_location():
    if not g.user or g.user['role'] != 'lecturer':
        return jsonify({'success': False, 'message': 'Not authorized'}), 403
    data = request.get_json()
    lat = data.get('lat')
    lon = data.get('lon')
    if lat is None or lon is None:
        return jsonify({'success': False, 'message': 'Latitude and longitude required'}), 400
    try:
        lat = float(lat)
        lon = float(lon)
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid coordinates'}), 400
    # Update or insert current location
    existing = query_db('SELECT id FROM locations WHERE lecturer_id = ? AND name = ?', (g.user['id'], 'Current Location'), one=True)
    if existing:
        commit_db('UPDATE locations SET latitude = ?, longitude = ? WHERE id = ?', (lat, lon, existing['id']))
    else:
        commit_db('INSERT INTO locations (lecturer_id, name, latitude, longitude) VALUES (?,?,?,?)', (g.user['id'], 'Current Location', lat, lon))
    return jsonify({'success': True, 'message': 'Location updated successfully'})


@app.route('/favicon.ico')
def favicon():
    favicon_data = get_setting('favicon')
    if favicon_data and favicon_data.startswith('data:'):
        try:
            header, b64 = favicon_data.split(',', 1)
            mime = header.split(';')[0].split(':')[1]
            img_bytes = base64.b64decode(b64)
            return send_file(BytesIO(img_bytes), mimetype=mime)
        except Exception:
            pass
    return '', 204

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('error.html'), 500

@app.route('/health-status')
def health_status_page():
    """Visual health dashboard for system status monitoring."""
    db_status = "STABLE"
    db_connected = True
    try:
        db = get_db()
        with db.cursor() as cur:
            cur.execute('SELECT 1')
    except Exception:
        db_status = "UNSTABLE"
        db_connected = False

    return render_template('health.html',
                           db_status=db_status,
                           db_connected=db_connected,
                           timestamp=datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC'))


# CLI: init DB on startup
if __name__ == '__main__':
    # Ensure database initialization runs inside the app context
    with app.app_context():
        # Optional: seed a super-admin from environment variables if provided
        try:
            admin_count_res = query_db("SELECT COUNT(*) as c FROM users WHERE role = 'admin'", (), one=True)
            admin_count = admin_count_res['c'] if admin_count_res else 0
        except Exception:
            admin_count = 0
        seed_user = os.environ.get('SUPERADMIN_SEED_USER')
        seed_pass = os.environ.get('SUPERADMIN_SEED_PASS')
        if seed_user and seed_pass and admin_count == 0:
            print('Seeding super-admin from environment...')
            try:
                commit_db('INSERT INTO users (role, username, password_hash, first_name, active) VALUES (?,?,?,?,?)',
                          ('admin', seed_user, generate_password_hash(seed_pass), os.environ.get('SUPERADMIN_SEED_NAME','Super'), 1))
                print('Super-admin created:', seed_user)
            except Exception as e:
                print('Failed to seed super-admin:', e)
    debug = os.environ.get('FLASK_DEBUG','0') == '1'
    host = os.environ.get('HOST','0.0.0.0')
    port = int(os.environ.get('PORT','5000'))

    # Enhanced port binding logic to handle occupied ports in local development
    import socket
    def is_port_in_use(p):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            return s.connect_ex(('localhost', p)) == 0

    if is_port_in_use(port) and not os.environ.get('PORT'):
        # Only try to find another port if PORT was not explicitly set in environment
        original_port = port
        while is_port_in_use(port):
            port += 1
            if port > original_port + 10:
                break
        print(f"Port {original_port} in use, switching to {port}")

    app.run(host=host, port=port, debug=debug)

# Vercel serverless handler
try:
    from serverless_wsgi import handle_request
    def handler(event, context):
        return handle_request(app, event, context)
except ImportError:
    pass
